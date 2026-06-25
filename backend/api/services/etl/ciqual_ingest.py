"""CIQUAL 2025 ETL ingest (CIQUAL-INGEST, 2026-06-26).

Reads [`backend/raw_ciqual/Table Ciqual 2025_ENG_2025_11_03.xlsx`](backend/raw_ciqual/Table%20Ciqual%202025_ENG_2025_11_03.xlsx)
and appends ~3,484 French foods (ANSES Centre d'Information sur la
Qualité des Aliments, 2025 release, English translation) into the
in-memory CNF DataFrames, using:

  - FoodIDs offset by 900,000 (CNF max ~503,381; WAFCT 700k; FDC 800k-825k → ~75k clear)
  - FoodGroupIDs 300-311 for the 12 CIQUAL top-level groups
  - FoodSourceID = 102 (CNF max 38, WAFCT 100, FDC 101)
  - NutrientSourceID = 10003 (CNF max 110, WAFCT 9999, FDC 10000-10002)
  - `source = 'ciqual'` on the new food_name_df rows
  - `data_type = 'ciqual_2025'`
  - `ciqual_code` carried into FoodCode as `CIQUAL_{alim_code}` for future
    cross-reference to the Agribalyse LCA catalogue (Agribalyse rows
    carry the same Ciqual code).

INFOODS → CNF NutrientID bridge: CIQUAL ships an explicit
`INFOODS codes` sheet mapping each nutrient column to an INFOODS tag.
The bridge reuses the same WAFCT-EXTEND IDENTICAL_TAGS + WAFCT_TO_CNF_TAG
+ WAFCT_TO_CNF_NUTRIENT_ID maps to convert those INFOODS tags to CNF
NutrientIDs — INFOODS is the same controlled vocabulary across both
sources, so the bridge code is reused as-is. ~60 of CIQUAL's 75 nutrients
bridge directly; the remainder (chloride, polyols, jones-factor metadata)
are dropped in v1.

Called once per process from [`api.cnf_cache.get_api_cnf_pipeline()`](backend/api/cnf_cache.py)
on first pipeline access, after the FDC ingest. Graceful degrade: if the
workbook is missing, the ingest is a no-op and the platform runs without
CIQUAL.

Performance: ~3 s wall-clock cold (openpyxl read + DataFrame construction).
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

# Reuse the WAFCT INFOODS bridge layers — same controlled vocabulary, same
# CNF target. Keeps the bridge logic in one place; CIQUAL just supplies a
# different per-column INFOODS-tag assignment.
from api.services.etl.wafct_ingest import (
    IDENTICAL_TAGS,
    WAFCT_TO_CNF_TAG,
    WAFCT_TO_CNF_NUTRIENT_ID,
)

logger = logging.getLogger(__name__)


# --- Constants (allocations are deliberately distant from CNF/WAFCT/FDC) --

CIQUAL_FOOD_ID_OFFSET    = 900_000   # CNF max 503,381; WAFCT 700k; FDC 800-825k
CIQUAL_FOOD_GROUP_BASE   = 300       # CNF 1-25, WAFCT 50-69, FDC 70-264 — 300 leaves clear headroom
CIQUAL_FOOD_SOURCE_ID    = 102       # CNF max 38, WAFCT 100, FDC 101
CIQUAL_NUTRIENT_SOURCE_ID = 10003    # CNF max 110, WAFCT 9999, FDC 10000-10002
CIQUAL_COUNTRY_CODE      = 'FR'

# Subdirectory layout under backend/.
CIQUAL_SUBDIR = 'raw_ciqual'
CIQUAL_FILENAME = 'Table Ciqual 2025_ENG_2025_11_03.xlsx'

# Guard rail: if CNF ever grows past this, raise our offset.
CNF_MAX_FOOD_ID_GUARD = 890_000


# --- Result payload ------------------------------------------------------

@dataclass
class CIQUALIngestResult:
    food_name_rows:        pd.DataFrame
    nutrient_amount_rows:  pd.DataFrame
    food_group_rows:       pd.DataFrame
    food_source_row:       pd.DataFrame
    nutrient_source_row:   pd.DataFrame
    bridge:                Dict[str, int] = field(default_factory=dict)
    unmapped_tags:         List[str]      = field(default_factory=list)
    stats:                 Dict[str, Any] = field(default_factory=dict)

    @property
    def food_count(self) -> int:
        return len(self.food_name_rows)


# --- Path resolution ------------------------------------------------------

def _ciqual_path() -> Path:
    """Resolve the CIQUAL workbook path."""
    try:
        from django.conf import settings
        candidate = Path(settings.BASE_DIR) / CIQUAL_SUBDIR / CIQUAL_FILENAME
        if candidate.exists():
            return candidate
    except Exception:  # noqa: BLE001
        pass
    here = Path(__file__).resolve()
    return here.parents[3] / CIQUAL_SUBDIR / CIQUAL_FILENAME


def workbook_present() -> bool:
    """True if CIQUAL XLSX is reachable. Lets callers graceful-degrade."""
    return _ciqual_path().exists()


# --- INFOODS sheet → per-column tag assignment ----------------------------

def _build_infoods_column_assignment(
    workbook_path: Path,
) -> Dict[int, str]:
    """Map column-index-in-food-composition → INFOODS tag, using the
    `INFOODS codes` sheet's `ORIGCPCD` → `INFDSTAG` mapping plus the
    column-name lookup from `food composition` row 1.

    The INFOODS sheet names the original CIQUAL code (e.g. 10260 = Iron,
    INFOODS tag FE), and the food-composition header contains the same
    English nutrient name. We match by header substring rather than by
    code so the column index can be discovered safely.

    Returns: `{column_index: infoods_tag}` for every column that has a
    matchable INFOODS tag. Columns without a tag (e.g. Jones factor,
    metadata) are absent from the dict.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws_food = wb['food composition']
    header = next(ws_food.iter_rows(min_row=1, max_row=1, values_only=True))
    header_norm = [(_norm(h)) for h in header]

    ws_inf = wb['INFOODS codes']
    inf_rows = list(ws_inf.iter_rows(min_row=2, values_only=True))
    # The INFOODS sheet has rows like (tag, code, name). Match `name`
    # against the food-composition header (substring-tolerant — units
    # vary in punctuation).
    out: Dict[int, str] = {}
    for r in inf_rows:
        if not r or r[0] is None:
            continue
        tag = str(r[0]).strip()
        const_name = _norm(str(r[2] or ''))
        if not const_name:
            continue
        matched_col: Optional[int] = None
        # Find the matching food-composition column by name.
        for col_idx, h_norm in enumerate(header_norm):
            if h_norm and h_norm == const_name:
                matched_col = col_idx
                break
        if matched_col is None:
            # If exact-match fails, try a loose contains match on the
            # part before the unit parenthesis.
            const_root = const_name.split('(')[0].strip()
            if not const_root:
                continue
            for col_idx, h_norm in enumerate(header_norm):
                if h_norm and const_root and const_root in h_norm and col_idx not in out:
                    matched_col = col_idx
                    break
        if matched_col is None:
            continue
        # Special-case ENERC: the CIQUAL INFOODS sheet uses bare "ENERC" for
        # both kJ and kcal rows. Disambiguate using the header text so the
        # bridge can map kJ → CNF NID 268 (ENERC_KJ) and kcal → CNF NID 208
        # (ENERC_KCAL) independently.
        if tag == 'ENERC':
            header_text = _norm(header[matched_col])
            if 'kcal' in header_text:
                tag = 'ENERC_kcal'
            elif 'kj' in header_text:
                tag = 'ENERC_kJ'
        out[matched_col] = tag
    return out


def _norm(s: Optional[str]) -> str:
    """Normalise a column-name or constituent-name string for matching:
    lower-case, collapse whitespace, strip slashes-and-units variability."""
    if s is None:
        return ''
    text = str(s).strip().lower()
    text = re.sub(r'\s+', ' ', text)
    text = text.replace('/100g', ' 100g').replace('/100 g', ' 100g')
    text = text.replace('µg', 'ug').replace('μg', 'ug')
    return text


# --- INFOODS tag → CNF NutrientID bridge ----------------------------------

def _build_infoods_to_cnf_bridge(cnf_pipeline) -> Dict[str, int]:
    """Reuse the WAFCT INFOODS→CNF bridge logic on the CNF pipeline.
    Same controlled vocabulary, same CNF target — CIQUAL just supplies
    per-column INFOODS tags from its own sheet.
    """
    nn = cnf_pipeline.nutrient_name_df
    cnf_tag_to_id: Dict[str, int] = {}
    for _, r in nn.iterrows():
        t = r.get('Tagname')
        if t is None or pd.isna(t):
            continue
        ts = str(t).strip()
        if ts and ts not in cnf_tag_to_id:
            cnf_tag_to_id[ts] = int(r['NutrientID'])

    bridge: Dict[str, int] = {}
    for tag in IDENTICAL_TAGS:
        nid = cnf_tag_to_id.get(tag)
        if nid is not None:
            bridge[tag] = nid
    for src_t, cnf_t in WAFCT_TO_CNF_TAG.items():
        nid = cnf_tag_to_id.get(cnf_t)
        if nid is not None:
            bridge[src_t] = nid
    for src_t, nid in WAFCT_TO_CNF_NUTRIENT_ID.items():
        bridge[src_t] = nid

    # CIQUAL-specific extensions: tags CIQUAL ships that the WAFCT bridge
    # doesn't already cover. Most have a direct CNF Tagname match; a few
    # need explicit alias mappings or direct NutrientID overrides.
    ciqual_extra_tag_to_cnf_tag = {
        # Core macros / energy
        'PROCNT':       'PROCNT',           # Protein (NID 203)
        'CHOAVL':       'CHOCDF',           # Available carb → CNF Carb-by-difference (NID 205)
        'SUGAR':        'SUGAR',            # Total sugars (NID 269)
        'STARCH':       'STARCH',           # Starch (NID 209)
        'ENERC_kcal':   'ENERC_KCAL',       # Energy kcal (NID 208) — disambiguated upstream
        'ENERC_kJ':     'ENERC_KJ',         # Energy kJ (NID 268)
        # Vitamins
        'CHOL-':        'CHOLE',            # Cholesterol
        'VITD-':        'VITD',             # Vitamin D (D2+D3)
        'VITE-':        'TOCPHA',           # Vitamin E → α-tocopherol
        'VITB6-':       'VITB6A',
        'FIB-':         'FIBTG',            # Total fibre
        'FOL':          'FOLAC',            # Total folate (closest; DFE handled via FOLDFE)
        'RAE':          'VITA_RAE',         # Vitamin A RAE
        # Minerals not in the WAFCT bridge
        'MN':           'MN',               # Manganese (NID 315)
        # Dropped explicitly (sub-decompositions CNF doesn't carry as separate columns)
        'POLYL':        None,               # Polyols — no CNF equivalent
        'OA':           None,               # Organic acids — no CNF equivalent
        'CLD':          None,               # Chloride — CNF only ships sodium
        'ERGCAL':       None,               # Vit D2 sub — combined as VITD-
        'CHOCAL':       None,               # Vit D3 sub — combined as VITD-
        'VITK1':        None, 'VITK2': None,  # CNF has combined NID 430
        # Sub-fatty acids CNF doesn't carry per-chain
        'F4D0':         None, 'F6D0':  None, 'F8D0':  None, 'F10D0': None,
        'F12D0':        None, 'F14D0': None, 'F16D0': None, 'F18D0': None,
        'F18D1CN9':     None, 'F20D4N6': None, 'F20D5N3': None, 'F22D6N3': None,
        # Sub-sugars CNF doesn't disaggregate
        'FRUS':         None, 'GALS': None, 'GLUS': None, 'LACS': None, 'MALS': None,
    }
    for src_t, cnf_t in ciqual_extra_tag_to_cnf_tag.items():
        if cnf_t is None:
            continue
        nid = cnf_tag_to_id.get(cnf_t)
        if nid is not None:
            bridge[src_t] = nid

    # Direct CNF NutrientID overrides for tags CNF lacks in its Tagname column.
    # Mirrors the WAFCT_TO_CNF_NUTRIENT_ID pattern.
    ciqual_extra_tag_to_cnf_nid = {
        'SE':       317,    # Selenium (CNF NID 317; Tagname column often null)
        'ID':       314,    # Iodine
        'PANTAC':   410,    # Pantothenic acid (NID 410)
    }
    for src_t, nid in ciqual_extra_tag_to_cnf_nid.items():
        bridge[src_t] = nid

    return bridge


# --- Workbook reader ------------------------------------------------------

def _load_ciqual_food_composition(
    workbook_path: Path,
    col_to_tag: Dict[int, str],
) -> List[Dict[str, Any]]:
    """Read the `food composition` sheet into a list of dicts, one per food."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb['food composition']
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    _header = next(rows_iter)  # skip header (we already used it to build col_to_tag)
    out: List[Dict[str, Any]] = []
    for row in rows_iter:
        if not row or row[6] is None:
            continue
        try:
            alim_code = int(row[6])
        except (TypeError, ValueError):
            continue
        grp_code = str(row[0] or '').strip()
        grp_name = str(row[3] or '').strip()
        name_en  = str(row[7] or '').strip()
        name_sci = str(row[8] or '').strip()
        nutrients: Dict[str, Optional[float]] = {}
        for col_idx, tag in col_to_tag.items():
            if col_idx >= len(row):
                continue
            v = _parse_value(row[col_idx])
            if v is not None:
                # If the same tag is supplied by multiple columns (e.g.
                # ENERC in kJ + kcal), prefer kcal for CNF compatibility:
                # take the first one matched then skip duplicates.
                if tag not in nutrients:
                    nutrients[tag] = v
        out.append({
            'alim_code':   alim_code,
            'group_code':  grp_code,
            'group_name':  grp_name,
            'name_en':     name_en,
            'name_sci':    name_sci,
            'nutrients':   nutrients,
        })
    return out


def _parse_value(v: Any) -> Optional[float]:
    """Parse a CIQUAL nutrient cell. Handles:
      - blank / dash (returned as None)
      - '<X' (limit-of-quantification; treat as zero)
      - 'traces' / 'tr' (returned as zero)
      - numeric (returned as float)
      - bracketed values like '[10.6]' (analytical-method-specific; strip brackets)
    """
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in ('-', '–'):
        return None
    if s.lower() in ('traces', 'tr'):
        return 0.0
    if s.startswith('<'):
        return 0.0  # below limit of quantification
    if s.startswith('[') and s.endswith(']'):
        s = s[1:-1].strip()
    # Strip commas used as decimal separators (rare in EN release; mostly . used).
    s = s.replace(',', '.')
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


# --- Main entry point -----------------------------------------------------

def ingest_ciqual(cnf_pipeline) -> CIQUALIngestResult:
    """Read CIQUAL 2025 XLSX, build DataFrames keyed for append to CNF schema."""
    path = _ciqual_path()
    if not path.exists():
        raise FileNotFoundError(f'CIQUAL workbook not found at {path}')

    cnf_max = int(cnf_pipeline.food_name_df['FoodID'].max())
    if cnf_max >= CNF_MAX_FOOD_ID_GUARD:
        raise RuntimeError(
            f'CNF max FoodID ({cnf_max}) approaches the CIQUAL offset region. '
            f'Raise CIQUAL_FOOD_ID_OFFSET above {cnf_max + 50_000} before continuing.'
        )

    logger.info('CIQUAL ingest: reading %s', path)

    col_to_tag = _build_infoods_column_assignment(path)
    logger.info('CIQUAL ingest: matched %d food-composition columns to INFOODS tags', len(col_to_tag))

    bridge = _build_infoods_to_cnf_bridge(cnf_pipeline)
    logger.info('CIQUAL ingest: %d INFOODS->CNF NutrientID mappings', len(bridge))

    foods = _load_ciqual_food_composition(path, col_to_tag)
    logger.info('CIQUAL ingest: loaded %d food rows', len(foods))

    # 1. food_group_rows — discover top-level groups from the data.
    seen_groups: Dict[str, str] = {}
    for f in foods:
        if f['group_code'] and f['group_code'] not in seen_groups:
            seen_groups[f['group_code']] = f['group_name']
    group_codes_sorted = sorted(seen_groups.keys())
    fg_rows = []
    fg_id_by_code: Dict[str, int] = {}
    for i, code in enumerate(group_codes_sorted):
        fg_id = CIQUAL_FOOD_GROUP_BASE + i
        fg_id_by_code[code] = fg_id
        fg_rows.append({
            'FoodGroupID':    fg_id,
            'FoodGroupCode':  f'CIQUAL_{code}',
            'FoodGroupName':  f'CIQUAL — {seen_groups[code] or "Group " + code}',
            'FoodGroupNameF': '',
        })
    food_group_df = pd.DataFrame(fg_rows)

    # 2. food_source_row + nutrient_source_row
    food_source_row = pd.DataFrame([{
        'FoodSourceID':           CIQUAL_FOOD_SOURCE_ID,
        'FoodSourceCode':         'CIQUAL_2025',
        'FoodSourceDescription':  'ANSES Centre d\'Information sur la Qualité des Aliments (CIQUAL) 2025, English release',
        'FoodSourceDescriptionF': 'Table Ciqual 2025 — ANSES, version anglaise',
    }])
    nutrient_source_row = pd.DataFrame([{
        'NutrientSourceID':           CIQUAL_NUTRIENT_SOURCE_ID,
        'NutrientSourceCode':         'CIQUAL_2025',
        'NutrientSourceDescription':  'Ingested from ANSES CIQUAL 2025 (EN)',
        'NutrientSourceDescriptionF': 'Importé depuis ANSES CIQUAL 2025 (EN)',
    }])

    # 3. food_name_rows + nutrient_amount_rows
    today = datetime.today()
    food_name_records = []
    nutrient_amount_records = []
    # Sort by alim_code for stable FoodID allocation.
    foods_sorted = sorted(foods, key=lambda f: f['alim_code'])
    for idx, food in enumerate(foods_sorted):
        food_id = CIQUAL_FOOD_ID_OFFSET + idx
        fg_id = fg_id_by_code.get(food['group_code'])
        if fg_id is None:
            logger.warning('CIQUAL food alim_code=%s has unknown group_code %s — skipping',
                           food['alim_code'], food['group_code'])
            continue
        food_name_records.append({
            'FoodID':                food_id,
            'FoodCode':              f'CIQUAL_{food["alim_code"]}',
            'FoodGroupID':           fg_id,
            'FoodSourceID':          CIQUAL_FOOD_SOURCE_ID,
            'FoodDescription':       food['name_en'],
            'FoodDescriptionF':      '',
            'FoodDateOfEntry':       today,
            'FoodDateOfPublication': today,
            'CountryCode':           CIQUAL_COUNTRY_CODE,
            'ScientificName':        food['name_sci'],
            'source':                'ciqual',
            'data_type':             'ciqual_2025',
            'fdc_id':                pd.NA,
        })
        # Per-nutrient rows
        for tag, value in food['nutrients'].items():
            nid = bridge.get(tag)
            if nid is None:
                continue  # unmapped — silently drop in v1 (tracked in unmapped_tags)
            nutrient_amount_records.append({
                'FoodID':               food_id,
                'NutrientID':           nid,
                'NutrientValue':        value,
                'StandardError':        None,
                'NumberofObservations': None,
                'NutrientSourceID':     CIQUAL_NUTRIENT_SOURCE_ID,
                'NutrientDateOfEntry':  today,
            })

    food_name_df = pd.DataFrame(food_name_records)
    nutrient_amount_df = pd.DataFrame(nutrient_amount_records)

    encountered_tags = {t for f in foods for t in f['nutrients']}
    unmapped = sorted(encountered_tags - set(bridge.keys()))

    stats = {
        'foods_loaded':           len(foods),
        'foods_emitted':          len(food_name_df),
        'nutrient_rows_emitted':  len(nutrient_amount_df),
        'unique_tags_used':       len({r['NutrientID'] for r in nutrient_amount_records}),
        'bridge_size':            len(bridge),
        'unmapped_tag_count':     len(unmapped),
        'ciqual_food_id_min':     int(food_name_df['FoodID'].min()) if not food_name_df.empty else None,
        'ciqual_food_id_max':     int(food_name_df['FoodID'].max()) if not food_name_df.empty else None,
        'food_group_count':       len(food_group_df),
    }
    logger.info('CIQUAL ingest done: %s', stats)

    return CIQUALIngestResult(
        food_name_rows=food_name_df,
        nutrient_amount_rows=nutrient_amount_df,
        food_group_rows=food_group_df,
        food_source_row=food_source_row,
        nutrient_source_row=nutrient_source_row,
        bridge=bridge,
        unmapped_tags=unmapped,
        stats=stats,
    )


def append_to_pipeline(cnf_pipeline, result: CIQUALIngestResult) -> None:
    """Append CIQUAL rows in-place onto the live CNF pipeline DataFrames.
    Safe to call AT MOST ONCE per pipeline instance — the cache singleton
    in `api.cnf_cache` enforces this. Mirrors WAFCT/FDC append pattern.
    """
    p = cnf_pipeline
    if 'source' not in p.food_name_df.columns:
        p.food_name_df['source'] = 'cnf'
    if 'data_type' not in p.food_name_df.columns:
        p.food_name_df['data_type'] = None
    if 'fdc_id' not in p.food_name_df.columns:
        p.food_name_df['fdc_id'] = pd.NA

    existing_food_ids = set(p.food_name_df['FoodID'].dropna().astype(int).tolist())
    new_foods = result.food_name_rows[
        ~result.food_name_rows['FoodID'].isin(existing_food_ids)
    ]
    p.food_name_df = pd.concat([p.food_name_df, new_foods], ignore_index=True)

    existing_group_ids = set(p.food_group_df['FoodGroupID'].dropna().astype(int).tolist())
    new_groups = result.food_group_rows[
        ~result.food_group_rows['FoodGroupID'].isin(existing_group_ids)
    ]
    p.food_group_df = pd.concat([p.food_group_df, new_groups], ignore_index=True)

    existing_fs_ids = set(p.food_source_df['FoodSourceID'].dropna().astype(int).tolist())
    new_fs = result.food_source_row[
        ~result.food_source_row['FoodSourceID'].isin(existing_fs_ids)
    ]
    p.food_source_df = pd.concat([p.food_source_df, new_fs], ignore_index=True)

    existing_ns_ids = set(p.nutrient_source_df['NutrientSourceID'].dropna().astype(int).tolist())
    new_ns = result.nutrient_source_row[
        ~result.nutrient_source_row['NutrientSourceID'].isin(existing_ns_ids)
    ]
    p.nutrient_source_df = pd.concat([p.nutrient_source_df, new_ns], ignore_index=True)

    p.nutrient_amount_df = pd.concat(
        [p.nutrient_amount_df, result.nutrient_amount_rows], ignore_index=True,
    )

    p.nutrients_by_food = p._build_nutrients_by_food_index()
    logger.info(
        'CIQUAL append complete: %d total foods (+%d CIQUAL at IDs %d-%d), %d total nutrient rows',
        len(p.food_name_df), len(new_foods),
        int(new_foods['FoodID'].min()) if not new_foods.empty else 0,
        int(new_foods['FoodID'].max()) if not new_foods.empty else 0,
        len(p.nutrient_amount_df),
    )
