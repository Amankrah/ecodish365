"""WAFCT 2019 structure-discovery harness (WAFCT-EXPLORE Phase 1+2, 2026-05-24).

Read-only inspector for the West African Food Composition Table (WAFCT) 2019
workbook ([`backend/raw_wafct/WAFCT_2019.xlsx`](backend/raw_wafct/WAFCT_2019.xlsx)).
Reads all 12 sheets, dumps a structured JSON inventory + console summary, and
builds the curated **INFOODS tagname → CNF NutrientName** bridge that any
future integration will pivot on.

Pure openpyxl + standard library — no Django, no LLM, no network. Runs in
~5 s. Output:
  - `_explore_wafct_structure.json` — machine-readable schema + counts
  - stdout — skimmable per-sheet summary

Run from `backend/`:
    set PYTHONIOENCODING=utf-8 && python _explore_wafct.py
"""
from __future__ import annotations

import json
import os
import re
import sys
from collections import Counter, defaultdict
from typing import Any, Dict, List, Optional, Tuple

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:  # noqa: BLE001
    pass

import openpyxl  # noqa: E402

WAFCT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          'raw_wafct', 'WAFCT_2019.xlsx')
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        '_explore_wafct_structure.json')

# WAFCT food Codes look like "01_172" — first two chars name the food group.
CODE_RE = re.compile(r'^(\d{2})_\d+$')
# Banding rows have a literal "/" between English + French group names.
GROUP_BAND_RE = re.compile(r'^[^/]+/[^/]+$')


# --- Phase 2: curated INFOODS -> CNF NutrientName bridge ------------------
#
# The Phase 1 inspection of WAFCT sheet `02 Components` enumerated 59
# tagnames; this dict pairs the ones with a clear CNF NutrientName
# equivalent (verified against `backend/api/cnf_data_pipeline.py`'s
# `nutrients_by_food` keys, which come from CNF's `NUTRIENT_NAME.csv`).
#
# Two WAFCT-side facts the table absorbs:
#   - `ENERC` appears TWICE in the 39-set (once for kJ, once for kcal);
#     disambiguated by column-index, not tagname.
#   - Tags like "FAT or [FATCE]" are an INFOODS convention — bracketed
#     variant is method-specific (FATCE = continuous-flow-extraction). For
#     CNF bridging the unbracketed tag is the canonical one.
#
# Inventoried gaps (WAFCT-only nutrients with NO CNF counterpart):
#   - PHYTCPP / IP3 / IP4 / IP5 / IP6 (phytate + inositol phosphates —
#     anti-nutrients, regionally important in millet / sorghum)
#   - SOP (sum of proximates — a check column, not a nutrient)
#   - XFA / XN (Atwater conversion factors)
#   - EDIBLE1 / EDIBLE2 (edible-portion coefficients — metadata)
#
# Inventoried gaps (CNF-only nutrients not in WAFCT 39-set):
#   - Granular fatty-acid breakdown (CNF carries 30+ individual FAs)
#   - Sugar partitioning (CNF: TOTAL SUGARS, GLUCOSE, FRUCTOSE etc.)
#   - Several B vitamins beyond what WAFCT carries
#
INFOODS_TO_CNF: Dict[str, str] = {
    # Energy
    'ENERC_kJ':    'ENERGY (KILOJOULES)',
    'ENERC_kcal':  'ENERGY (KILOCALORIES)',
    # Proximate composition
    'WATER':       'MOISTURE',
    'PROTCNT':     'PROTEIN',
    'FAT':         'FAT (TOTAL LIPIDS)',
    'FATCE':       'FAT (TOTAL LIPIDS)',
    'CHOAVLDF':    'CARBOHYDRATE, TOTAL (BY DIFFERENCE)',
    'FIBTG':       'FIBRE, TOTAL DIETARY',
    'FIBC':        'FIBRE, CRUDE',
    'ALC':         'ALCOHOL',
    'ASH':         'ASH, TOTAL',
    # Minerals
    'CA':          'CALCIUM',
    'FE':          'IRON',
    'MG':          'MAGNESIUM',
    'P':           'PHOSPHORUS',
    'K':           'POTASSIUM',
    'NA':          'SODIUM',
    'ZN':          'ZINC',
    'CU':          'COPPER',
    # Vitamins
    'VITA':        'VITAMIN A',
    'VITA_RAE':    'RETINOL ACTIVITY EQUIVALENTS',
    'RETOL':       'RETINOL',
    'CARTBEQ':     'BETA CAROTENE',
    'CARTB':       'BETA CAROTENE',
    'CARTA':       'ALPHA CAROTENE',
    'CRYPXB':      'CRYPTOXANTHIN, BETA',
    'VITD':        'VITAMIN D (D2 + D3)',
    'VITE':        'VITAMIN E (ALPHA-TOCOPHEROL)',
    'TOCPHA':      'TOCOPHEROL, ALPHA',
    'TOCPHB':      'TOCOPHEROL, BETA',
    'TOCPHG':      'TOCOPHEROL, GAMMA',
    'TOCPHD':      'TOCOPHEROL, DELTA',
    'THIA':        'THIAMIN',
    'RIBF':        'RIBOFLAVIN',
    'NIA':         'NIACIN',
    'NIAEQ':       'NIACIN EQUIVALENT (NE)',
    'TRP':         'TRYPTOPHAN',
    'VITB6C':      'VITAMIN B-6',
    'FOL':         'FOLATE, TOTAL',
    'FOLSUM':      'FOLATE, TOTAL',
    'FOLAC':       'FOLIC ACID',
    'FOLFD':       'FOLATE, FOOD',
    'FOLDFE':      'FOLATE, DFE',
    'VITB12':      'VITAMIN B-12',
    'VITC':        'VITAMIN C',
    'CHOLE':       'CHOLESTEROL',
    # Fatty acids
    'FASAT':       'FATTY ACIDS, SATURATED, TOTAL',
    'FAMS':        'FATTY ACIDS, MONOUNSATURATED, TOTAL',
    'FAPU':        'FATTY ACIDS, POLYUNSATURATED, TOTAL',
    'F18D2CN6':    'FATTY ACIDS, n-6 LINOLEIC ACID, 18:2',
    'F18D3CN3':    'FATTY ACIDS, n-3 LINOLENIC ACID, 18:3',
}

# WAFCT-only tags (no CNF equivalent — surfaced as orphans in the report).
WAFCT_ONLY_TAGS = {
    'PHYTCPP', 'PHYTCP', 'IP3', 'IP4', 'IP5', 'IP6',
    'SOP', 'XFA', 'XN', 'EDIBLE1', 'EDIBLE2',
}


# --- Helpers --------------------------------------------------------------

def _safe_str(v: Any, max_len: int = 200) -> str:
    if v is None:
        return ''
    s = str(v).replace('\n', ' ').strip()
    return s[:max_len]


def _clean_tag(v: Any) -> str:
    """Strip a tagname cell down to its canonical form ('ENERC ' → 'ENERC').

    Some cells embed an 'or' alternative ('FAT or [FATCE]') or carry
    leading/trailing whitespace; this canonicalises to the unbracketed
    primary tag so the INFOODS_TO_CNF lookup is uniform.
    """
    if v is None:
        return ''
    s = str(v).strip()
    # "FAT or [FATCE]" → "FAT"
    if ' or ' in s:
        s = s.split(' or ', 1)[0].strip()
    # "[FOLSUM]" → "FOLSUM"
    s = s.strip('[]').strip()
    return s


def _is_banding_row(row: Tuple[Any, ...]) -> bool:
    """True if this row is a food-group header (Cereals/Céréales, etc.)
    rather than an actual food row."""
    if not row or row[0] is None:
        return False
    col0 = _safe_str(row[0])
    if not col0 or not GROUP_BAND_RE.match(col0):
        return False
    # All other cols should be empty
    return all((c is None or _safe_str(c) == '') for c in row[1:])


def _is_data_row(row: Tuple[Any, ...]) -> bool:
    """True if col 0 matches the WAFCT '\\d{2}_\\d+' Code pattern."""
    if not row or row[0] is None:
        return False
    return bool(CODE_RE.match(_safe_str(row[0])))


def _detect_mojibake(s: str) -> bool:
    """U+FFFD REPLACEMENT CHARACTER indicates encoding corruption in the
    SOURCE file (not just our console)."""
    return '�' in s


# --- Per-sheet inspectors -------------------------------------------------

def inspect_sheet_01_intro(wb) -> Dict[str, Any]:
    ws = wb['01 Introduction']
    out: Dict[str, Any] = {'lines_sampled': []}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 40:
            break
        for cell in row:
            if cell is not None:
                s = _safe_str(cell, max_len=300)
                if s:
                    out['lines_sampled'].append(s)
    out['n_lines_sampled'] = len(out['lines_sampled'])
    return out


def inspect_sheet_02_components(wb) -> Dict[str, Any]:
    """Build the full INFOODS dictionary from sheet 02."""
    ws = wb['02 Components']
    components: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row[0] is None:
            continue
        tag_raw = _safe_str(row[2])
        # Skip the secondary French-header row (row index 1) which echoes
        # the literal string "INFOODS tagname" into the tagname column.
        if not tag_raw or tag_raw == 'INFOODS tagname':
            continue
        components.append({
            'tag_raw':      tag_raw,
            'tag_canonical': _clean_tag(tag_raw),
            'name_en':      _safe_str(row[0]),
            'name_fr':      _safe_str(row[1]),
            'unit':         _safe_str(row[3]),
            'denominator':  _safe_str(row[4]),
            'sig_figs':     _safe_str(row[5]),
            'max_decimals': _safe_str(row[6]),
            'datasheet':    _safe_str(row[7]),
            'method_en':    _safe_str(row[8], max_len=400),
        })
    return {
        'n_components': len(components),
        'tagnames_canonical': sorted({c['tag_canonical'] for c in components}),
        'components': components,
    }


def _inspect_nutrient_table(
    wb, sheet_name: str, expected_nutrient_count: int,
) -> Dict[str, Any]:
    """Generic inspector for an NV_sum_NN / NV_stat_NN sheet."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    n_total_rows = len(rows)
    n_total_cols = len(rows[0]) if rows else 0

    # Row 0 = English headers, Row 1 = French, Row 2 = INFOODS tagnames.
    header_en = [_safe_str(c, max_len=80) for c in rows[0]]
    header_tags_canonical: List[str] = []
    if n_total_rows >= 3:
        for i, c in enumerate(rows[2]):
            tag = _clean_tag(c)
            header_tags_canonical.append(tag)

    # Nutrient columns start at index 5 (after Code/EN/FR/Sci/Biblio)
    # — but actually the EDIBLE1 column appears at idx 5 in our peek,
    # so nutrient values start at idx 5 in the 39-set sheet.
    nutrient_col_start = next(
        (i for i, t in enumerate(header_tags_canonical) if t == 'EDIBLE1'),
        5,
    )
    nutrient_tags = header_tags_canonical[nutrient_col_start:]
    nutrient_tags_clean = [t for t in nutrient_tags if t]

    # Walk data rows: distinguish banding rows from real foods.
    n_banding = 0
    n_data = 0
    n_blank = 0
    n_other = 0
    group_codes_seen: Counter = Counter()
    banding_labels: List[str] = []
    mojibake_lines: List[str] = []
    nutrient_nonnull_count: Dict[int, int] = defaultdict(int)
    sample_data_rows: List[Dict[str, Any]] = []

    for row in rows[3:]:
        if all((c is None or _safe_str(c) == '') for c in row):
            n_blank += 1
            continue
        if _is_banding_row(row):
            n_banding += 1
            label = _safe_str(row[0])
            banding_labels.append(label)
            if _detect_mojibake(label):
                mojibake_lines.append(label)
            continue
        if _is_data_row(row):
            n_data += 1
            code = _safe_str(row[0])
            m = CODE_RE.match(code)
            if m:
                group_codes_seen[m.group(1)] += 1
            # Per-nutrient non-null counter
            for col_idx in range(nutrient_col_start, n_total_cols):
                v = row[col_idx]
                if v is not None and _safe_str(v) != '':
                    nutrient_nonnull_count[col_idx] += 1
            # First 5 data rows: persist for memo sampling
            if len(sample_data_rows) < 5:
                sample_data_rows.append({
                    'code':  code,
                    'name_en': _safe_str(row[1]),
                    'name_fr': _safe_str(row[2]),
                    'biblio': _safe_str(row[4]),
                    'edible': _safe_str(row[5]),
                    'nutrients_first10': [
                        _safe_str(row[col_idx])
                        for col_idx in range(nutrient_col_start, min(nutrient_col_start + 10, n_total_cols))
                    ],
                })
            # Mojibake check on free-text fields
            for c in (row[1], row[2]):
                if c is not None and _detect_mojibake(_safe_str(c)):
                    mojibake_lines.append(_safe_str(c))
            continue
        n_other += 1

    # Coverage per nutrient = % of data rows with a non-null value
    nutrient_coverage: List[Dict[str, Any]] = []
    for col_idx in range(nutrient_col_start, n_total_cols):
        tag = header_tags_canonical[col_idx] if col_idx < len(header_tags_canonical) else ''
        name = header_en[col_idx] if col_idx < len(header_en) else ''
        nonnull = nutrient_nonnull_count.get(col_idx, 0)
        coverage = (nonnull / n_data) if n_data else 0.0
        nutrient_coverage.append({
            'col_idx':       col_idx,
            'tag_canonical': tag,
            'name_en':       name,
            'nonnull':       nonnull,
            'coverage_pct':  round(coverage * 100, 1),
        })

    return {
        'n_total_rows':      n_total_rows,
        'n_total_cols':      n_total_cols,
        'n_data_rows':       n_data,
        'n_banding_rows':    n_banding,
        'n_blank_rows':      n_blank,
        'n_other_rows':      n_other,
        'expected_nutrient_count': expected_nutrient_count,
        'nutrient_col_start': nutrient_col_start,
        'n_nutrient_cols':   len(nutrient_tags_clean),
        'nutrient_tags_clean': nutrient_tags_clean,
        'group_codes_seen':  dict(sorted(group_codes_seen.items())),
        'n_groups':          len(group_codes_seen),
        'banding_labels':    banding_labels,
        'mojibake_lines':    mojibake_lines[:20],
        'mojibake_count':    len(mojibake_lines),
        'nutrient_coverage': nutrient_coverage,
        'sample_data_rows':  sample_data_rows,
    }


def inspect_sheet_07_yield(wb) -> Dict[str, Any]:
    ws = wb['07 Yield factors, sing_ing']
    rows = list(ws.iter_rows(values_only=True))
    n_data = sum(1 for r in rows if _is_data_row(r))
    yield_factors: List[float] = []
    for r in rows:
        if _is_data_row(r) and r[3] is not None:
            try:
                yield_factors.append(float(r[3]))
            except (TypeError, ValueError):
                pass
    return {
        'n_total_rows':  len(rows),
        'n_data_rows':   n_data,
        'n_with_yield':  len(yield_factors),
        'min_yield':     round(min(yield_factors), 3) if yield_factors else None,
        'max_yield':     round(max(yield_factors), 3) if yield_factors else None,
        'mean_yield':    round(sum(yield_factors) / len(yield_factors), 3) if yield_factors else None,
    }


def inspect_sheet_08_retention(wb) -> Dict[str, Any]:
    ws = wb['08 Retention factors']
    rows = list(ws.iter_rows(values_only=True))
    header_en = [_safe_str(c, max_len=40) for c in rows[0]] if rows else []
    nutrient_cols_with_retention = [h for h in header_en[4:] if h]
    n_data = sum(1 for r in rows[3:] if r[0] is not None and not _is_banding_row(r))
    return {
        'n_total_rows': len(rows),
        'n_data_rows': n_data,
        'retention_nutrients': nutrient_cols_with_retention,
        'n_retention_nutrients': len(nutrient_cols_with_retention),
    }


def inspect_sheet_09_mixed(wb) -> Dict[str, Any]:
    """Mixed dishes — ingredient lists per recipe.

    Schema in the workbook: each recipe starts with a row that has an
    Observation # in col 0 and a Yield factor in col 5; subsequent rows
    list ingredients (WAFCT Code in col 1, ingredient name in col 2,
    weight in col 4) until the next Observation #.
    """
    ws = wb['09 Mixed dishes']
    rows = list(ws.iter_rows(values_only=True))
    n_total = len(rows)
    n_recipes = 0
    n_ingredients = 0
    ingredients_per_recipe: List[int] = []
    current_ings = 0
    sample_recipes: List[Dict[str, Any]] = []
    current_recipe: Optional[Dict[str, Any]] = None

    for r in rows[1:]:  # skip header
        if r[0] is not None and str(r[0]).strip().isdigit():
            # New observation / recipe starts
            if current_recipe is not None:
                ingredients_per_recipe.append(current_ings)
                if len(sample_recipes) < 3:
                    sample_recipes.append(current_recipe)
            n_recipes += 1
            current_ings = 0
            current_recipe = {
                'obs_num': _safe_str(r[0]),
                'code':    _safe_str(r[1]),
                'name_en': _safe_str(r[2]),
                'yield':   _safe_str(r[5]),
                'ingredients': [],
            }
        elif r[1] is not None and CODE_RE.match(_safe_str(r[1])):
            # Ingredient row
            n_ingredients += 1
            current_ings += 1
            if current_recipe is not None and len(current_recipe['ingredients']) < 8:
                current_recipe['ingredients'].append({
                    'code':    _safe_str(r[1]),
                    'name_en': _safe_str(r[2]),
                    'weight':  _safe_str(r[4]),
                })
    if current_recipe is not None:
        ingredients_per_recipe.append(current_ings)
        if len(sample_recipes) < 3:
            sample_recipes.append(current_recipe)

    return {
        'n_total_rows':   n_total,
        'n_recipes':      n_recipes,
        'n_ingredients':  n_ingredients,
        'avg_ings_per_recipe': round(sum(ingredients_per_recipe) / len(ingredients_per_recipe), 2) if ingredients_per_recipe else 0,
        'max_ings_per_recipe': max(ingredients_per_recipe) if ingredients_per_recipe else 0,
        'cross_referenced_by_wafct_code': True,  # confirmed by CODE_RE.match on ingredient col 1
        'sample_recipes': sample_recipes,
    }


def inspect_sheet_10_foodex2(wb) -> Dict[str, Any]:
    ws = wb['10 FoodEx2 codes']
    rows = list(ws.iter_rows(values_only=True))
    n_data = 0
    n_with_foodex2 = 0
    n_exact_match = 0
    for r in rows[3:]:
        if not r or r[0] is None:
            continue
        if not _is_data_row(r):
            continue
        n_data += 1
        foodex2 = _safe_str(r[5])
        if foodex2 and foodex2 != '-':
            n_with_foodex2 += 1
        exact = _safe_str(r[7])
        if exact and exact.lower().startswith('yes'):
            n_exact_match += 1
    return {
        'n_total_rows':         len(rows),
        'n_data_rows':          n_data,
        'n_with_foodex2_code':  n_with_foodex2,
        'foodex2_coverage_pct': round((n_with_foodex2 / n_data) * 100, 1) if n_data else 0,
        'n_exact_match':        n_exact_match,
        'exact_match_pct':      round((n_exact_match / n_data) * 100, 1) if n_data else 0,
    }


def inspect_sheet_11_version_map(wb) -> Dict[str, Any]:
    ws = wb['11 2012 vs 2019 names and codes']
    rows = list(ws.iter_rows(values_only=True))
    return {
        'n_total_rows': len(rows),
        'header': [_safe_str(c) for c in (rows[0] if rows else [])],
    }


def inspect_sheet_12_biblio(wb) -> Dict[str, Any]:
    ws = wb['12 Data sources with BiblioID']
    rows = list(ws.iter_rows(values_only=True))
    n_entries = 0
    samples: List[Dict[str, str]] = []
    for r in rows[1:]:
        if r[0] is None or not _safe_str(r[0]):
            continue
        n_entries += 1
        if len(samples) < 10:
            samples.append({
                'biblio_id': _safe_str(r[0]),
                'reference': _safe_str(r[1], max_len=200),
            })
    return {
        'n_total_rows': len(rows),
        'n_entries':    n_entries,
        'sample_first10': samples,
    }


# --- INFOODS ↔ CNF bridge -------------------------------------------------

def build_infoods_cnf_bridge(components_inv: Dict[str, Any]) -> Dict[str, Any]:
    """Cross-reference WAFCT INFOODS tagnames against the curated CNF bridge."""
    wafct_tags = set(components_inv['tagnames_canonical'])
    bridge: List[Dict[str, Any]] = []
    matched = orphans_wafct = 0
    for tag in sorted(wafct_tags):
        # ENERC is canonically a single tag but represents both energy
        # units (kJ + kcal). The INFOODS_TO_CNF dict uses suffixed keys
        # (ENERC_kJ, ENERC_kcal) since the per-unit CNF NutrientName
        # differs — record both mappings.
        if tag == 'ENERC':
            bridge.append({
                'infoods_tag':   'ENERC (kJ)',
                'cnf_nutrient':  INFOODS_TO_CNF['ENERC_kJ'],
                'status':        'mapped',
            })
            bridge.append({
                'infoods_tag':   'ENERC (kcal)',
                'cnf_nutrient':  INFOODS_TO_CNF['ENERC_kcal'],
                'status':        'mapped',
            })
            matched += 2
            continue
        if tag in INFOODS_TO_CNF:
            bridge.append({
                'infoods_tag':   tag,
                'cnf_nutrient':  INFOODS_TO_CNF[tag],
                'status':        'mapped',
            })
            matched += 1
        elif tag in WAFCT_ONLY_TAGS:
            bridge.append({
                'infoods_tag':   tag,
                'cnf_nutrient':  None,
                'status':        'wafct_only_known',
            })
            orphans_wafct += 1
        else:
            bridge.append({
                'infoods_tag':   tag,
                'cnf_nutrient':  None,
                'status':        'unmapped_review',
            })
            orphans_wafct += 1
    return {
        'n_wafct_tags':   len(wafct_tags),
        'n_mapped':       matched,
        'n_unmapped':     orphans_wafct,
        'coverage_pct':   round((matched / len(wafct_tags)) * 100, 1) if wafct_tags else 0,
        'bridge':         bridge,
    }


# --- Reporter -------------------------------------------------------------

def _print_section(title: str) -> None:
    print('\n' + '=' * 90)
    print(title)
    print('=' * 90)


def main() -> int:
    print(f'WAFCT 2019 structure inspector — reading {WAFCT_PATH}')
    if not os.path.exists(WAFCT_PATH):
        print(f'ERROR: file not found at {WAFCT_PATH}', file=sys.stderr)
        return 1
    wb = openpyxl.load_workbook(WAFCT_PATH, read_only=True, data_only=True)
    print(f'  Loaded {len(wb.sheetnames)} sheets')

    report: Dict[str, Any] = {
        'harness':      'WAFCT-EXPLORE Phase 1+2 (2026-05-24)',
        'wafct_path':   WAFCT_PATH,
        'sheet_names':  wb.sheetnames,
    }

    _print_section('Sheet 01 Introduction')
    report['sheet_01_intro'] = inspect_sheet_01_intro(wb)
    print(f'  Sampled {report["sheet_01_intro"]["n_lines_sampled"]} text lines from intro/notes')
    for line in report['sheet_01_intro']['lines_sampled'][:8]:
        print(f'  > {line[:140]}')

    _print_section('Sheet 02 Components (INFOODS tagname dictionary)')
    report['sheet_02_components'] = inspect_sheet_02_components(wb)
    print(f'  {report["sheet_02_components"]["n_components"]} components defined')
    print(f'  {len(report["sheet_02_components"]["tagnames_canonical"])} unique canonical tagnames')

    _print_section('Sheet 03 NV_sum_39 (per 100g EP) — primary 39-nutrient table')
    report['sheet_03_nv_sum_39'] = _inspect_nutrient_table(wb, '03 NV_sum_39 (per 100g EP)', 39)
    s = report['sheet_03_nv_sum_39']
    print(f'  {s["n_data_rows"]} food rows, {s["n_banding_rows"]} group banding rows, {s["n_blank_rows"]} blank')
    print(f'  {s["n_nutrient_cols"]} nutrient columns (expected {s["expected_nutrient_count"]})')
    print(f'  {s["n_groups"]} food groups: {list(s["group_codes_seen"].items())}')
    print(f'  Banding labels:')
    for lbl in s['banding_labels']:
        print(f'    - {lbl}')
    print(f'  Mojibake (source-file encoding artefacts): {s["mojibake_count"]} lines')
    if s['mojibake_count']:
        for ln in s['mojibake_lines'][:3]:
            print(f'    [!] {ln[:80]}')
    print(f'  Per-nutrient coverage (top 12 by coverage):')
    for n in sorted(s['nutrient_coverage'], key=lambda x: -x['coverage_pct'])[:12]:
        print(f'    {n["tag_canonical"]:<14s}  {n["coverage_pct"]:>5.1f}%  ({n["nonnull"]:>4} / {s["n_data_rows"]})  {n["name_en"][:50]}')
    print(f'  Per-nutrient coverage (worst 5):')
    for n in sorted(s['nutrient_coverage'], key=lambda x: x['coverage_pct'])[:5]:
        print(f'    {n["tag_canonical"]:<14s}  {n["coverage_pct"]:>5.1f}%  ({n["nonnull"]:>4} / {s["n_data_rows"]})  {n["name_en"][:50]}')

    _print_section('Sheet 04 NV_stat_39 — statistics for 39-set')
    report['sheet_04_nv_stat_39'] = _inspect_nutrient_table(wb, '04 NV_stat_39 (per 100g EP)', 39)
    print(f'  {report["sheet_04_nv_stat_39"]["n_data_rows"]} stat rows (compare to {s["n_data_rows"]} sum rows)')

    _print_section('Sheet 05 NV_sum_57 — extended 57-nutrient table')
    report['sheet_05_nv_sum_57'] = _inspect_nutrient_table(wb, '05 NV_sum_57 (per 100g EP)', 57)
    s57 = report['sheet_05_nv_sum_57']
    print(f'  {s57["n_data_rows"]} food rows; {s57["n_nutrient_cols"]} nutrient cols')
    # Additional 57-set nutrients vs 39-set
    extra_in_57 = sorted(set(s57['nutrient_tags_clean']) - set(s['nutrient_tags_clean']))
    print(f'  Additional in 57-set (not in 39-set): {len(extra_in_57)} tags')
    print(f'    {extra_in_57}')

    _print_section('Sheet 06 NV_stat_57')
    report['sheet_06_nv_stat_57'] = _inspect_nutrient_table(wb, '06 NV_stat_57 (per 100g EP)', 57)
    print(f'  {report["sheet_06_nv_stat_57"]["n_data_rows"]} stat rows')

    _print_section('Sheet 07 Yield factors, single ingredients')
    report['sheet_07_yield'] = inspect_sheet_07_yield(wb)
    y = report['sheet_07_yield']
    print(f'  {y["n_with_yield"]} foods with yield factor (range [{y["min_yield"]}, {y["max_yield"]}], mean {y["mean_yield"]})')

    _print_section('Sheet 08 Retention factors')
    report['sheet_08_retention'] = inspect_sheet_08_retention(wb)
    r8 = report['sheet_08_retention']
    print(f'  {r8["n_data_rows"]} food/processing combos; retention for {r8["n_retention_nutrients"]} nutrients:')
    print(f'    {r8["retention_nutrients"]}')

    _print_section('Sheet 09 Mixed dishes (composite recipes)')
    report['sheet_09_mixed'] = inspect_sheet_09_mixed(wb)
    m = report['sheet_09_mixed']
    print(f'  {m["n_recipes"]} recipes, {m["n_ingredients"]} ingredient rows')
    print(f'  Avg {m["avg_ings_per_recipe"]} ings/recipe, max {m["max_ings_per_recipe"]}')
    print(f'  Cross-referenced by WAFCT Code: {m["cross_referenced_by_wafct_code"]}')
    if m['sample_recipes']:
        print(f'  Sample recipe #1: {m["sample_recipes"][0]["name_en"][:60]}')
        for ing in m['sample_recipes'][0]['ingredients'][:5]:
            print(f'    - {ing["weight"]:>6s}  {ing["code"]:<10s}  {ing["name_en"][:60]}')

    _print_section('Sheet 10 FoodEx2 codes (EFSA classification)')
    report['sheet_10_foodex2'] = inspect_sheet_10_foodex2(wb)
    fx = report['sheet_10_foodex2']
    print(f'  {fx["n_with_foodex2_code"]} / {fx["n_data_rows"]} foods have a FoodEx2 code ({fx["foodex2_coverage_pct"]}%)')
    print(f'  {fx["n_exact_match"]} marked Exact Match=Yes ({fx["exact_match_pct"]}%)')

    _print_section('Sheet 11 2012 vs 2019 name/code mapping')
    report['sheet_11_version_map'] = inspect_sheet_11_version_map(wb)
    print(f'  {report["sheet_11_version_map"]["n_total_rows"]} mapping rows')

    _print_section('Sheet 12 Data sources (bibliography)')
    report['sheet_12_biblio'] = inspect_sheet_12_biblio(wb)
    b = report['sheet_12_biblio']
    print(f'  {b["n_entries"]} bibliographic entries')
    print(f'  First 5:')
    for entry in b['sample_first10'][:5]:
        print(f'    {entry["biblio_id"]:<8s}  {entry["reference"][:90]}')

    _print_section('Phase 2: INFOODS → CNF NutrientName bridge')
    report['infoods_cnf_bridge'] = build_infoods_cnf_bridge(report['sheet_02_components'])
    br = report['infoods_cnf_bridge']
    print(f'  {br["n_wafct_tags"]} WAFCT tagnames total')
    print(f'  {br["n_mapped"]} mapped to CNF NutrientName ({br["coverage_pct"]}%)')
    print(f'  {br["n_unmapped"]} unmapped (WAFCT-only / under-review)')
    print()
    print('  Mapped bridge (first 20):')
    for entry in [e for e in br['bridge'] if e['status'] == 'mapped'][:20]:
        print(f'    {entry["infoods_tag"]:<14s}  ->  {entry["cnf_nutrient"]}')
    print()
    print('  Unmapped tagnames:')
    for entry in [e for e in br['bridge'] if e['status'] != 'mapped']:
        print(f'    {entry["infoods_tag"]:<14s}  ({entry["status"]})')

    # ---- Write JSON ------------------------------------------------------
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    _print_section('Output')
    print(f'  Wrote {OUT_PATH}')
    print(f'  File size: {os.path.getsize(OUT_PATH) / 1024:.1f} KB')

    return 0


if __name__ == '__main__':
    sys.exit(main())
