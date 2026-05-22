"""Build ReCiPe 2016 v1.1 factor packs from the official RIVM workbooks.

Idempotent ETL: 3 input workbooks -> 3 JSON packs + 1 meta JSON. Source-of-truth
for endpoint characterisation factors, normalisation scores, and country-specific
spatial CFs consumed by `methodology_factors.py` and `life_cycle_assessment.py`.

Run from `backend/` after activating the venv:

    python -m environmental_impact_model.etl.build_recipe2016_factor_packs

Use `--dry-run` to print row counts + sample rows without writing.

Outputs (in `backend/environmental_impact_model/data/`):

    recipe2016_endpoint_factors.json     # mid->endpoint conv factors per I/H/E
    recipe2016_normalization.json        # midpoint + endpoint norms per I/H/E
    recipe2016_country_factors.json      # per-country spatial CFs (5 categories)
    recipe2016_factor_packs_meta.json    # provenance + checksums

Schema versions are pinned and checked by the loader at runtime.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import subprocess
import sys
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# File paths (relative to this module's parent `environmental_impact_model/`)
# ---------------------------------------------------------------------------
DATA_DIR_REL = "environmental_impact_model/data"
CF_WORKBOOK_NAME = "ReCiPe2016_CFs_v1.1_20180117.xlsx"
NORM_WORKBOOK_NAME = "Normalization scores ReCiPe2016v1.1_20190514.xlsx"
COUNTRY_WORKBOOK_NAME = "ReCiPe2016_country factors_v1.1_20171221.xlsx"

# Output filenames
ENDPOINT_PACK_NAME = "recipe2016_endpoint_factors.json"
NORM_PACK_NAME = "recipe2016_normalization.json"
COUNTRY_PACK_NAME = "recipe2016_country_factors.json"
META_NAME = "recipe2016_factor_packs_meta.json"
ISO3_MAP_NAME = "country_iso3_map.json"  # written alongside

# Schema version — bump on backward-incompatible changes
SCHEMA_VERSION = "1.0"
METHODOLOGY = "recipe2016"
METHODOLOGY_VERSION = "v1.1"

PERSPECTIVES = ("I", "H", "E")
PERSPECTIVE_FULL_NAME = {
    "I": "Individualist",
    "H": "Hierarchist",
    "E": "Egalitarian",
}


# ---------------------------------------------------------------------------
# Workbook label -> internal code key for midpoint-to-endpoint factors.
# The first column of the 'Midpoint to endpoint factors' sheet uses verbose
# labels (with a typo: "Ionzing Radiation"). These map onto the canonical
# code keys consumed by `life_cycle_assessment._endpoint_from_midpoint_vector`.
# ---------------------------------------------------------------------------
ENDPOINT_LABEL_TO_KEY: Dict[str, str] = {
    # Human Health
    "Global Warming - Human health":                            "climate_change_human",
    "Stratospheric ozone depletion - Human health":             "ozone_depletion_human",
    "Ionzing Radiation - Human health":                         "ionizing_radiation_human",
    "Fine particulate matter formation - Human health":         "particulate_matter_human",
    "Photochemical ozone formation - Human health":             "photochemical_ozone_human",
    "Toxicity - Human health (cancer)":                         "human_toxicity_cancer",
    "Toxicity - Human health (non-cancer)":                     "human_toxicity_non_cancer",
    "Water consumption - human health":                         "water_use_human",
    # Terrestrial ecosystems
    "Global Warming - Terrestrial ecosystems":                  "climate_change_ecosystem",
    "Photochemical ozone formation - Terrestrial ecosystems":   "photochemical_ozone_ecosystem",
    "Acidification - Terrestrial ecosystems":                   "terrestrial_acidification_ecosystem",
    "Toxicity - Terrestrial ecosystems":                        "terrestrial_ecotoxicity_ecosystem",
    "Water consumption - terrestrial ecosystems":               "water_use_ecosystem_terrestrial",
    "Land use - occupation and transformation":                 "land_use_ecosystem",
    # Freshwater ecosystems
    "Global Warming - Freshwater ecosystems":                   "climate_change_ecosystem_freshwater",
    "Eutrophication - Freshwater ecosystems":                   "freshwater_eutrophication_ecosystem",
    "Toxicity - Freshwater ecosystems":                         "freshwater_ecotoxicity_ecosystem",
    "Water consumption -aquatic ecosystems":                    "water_use_ecosystem_freshwater",
    # Marine ecosystems
    "Toxicity - Marine ecosystems":                             "marine_ecotoxicity_ecosystem",
    "Eutrophication - Marine ecosystems":                       "marine_eutrophication_ecosystem",
    # Resources
    "Mineral resource scarcity":                                "mineral_scarcity",
    "Crude oil":                                                "fossil_scarcity_crude_oil",
    "Hard coal":                                                "fossil_scarcity_hard_coal",
    "Natural gas":                                              "fossil_scarcity_natural_gas",
    "Brown coal":                                               "fossil_scarcity_brown_coal",
    "Peat":                                                     "fossil_scarcity_peat",
}

# Midpoint normalisation row labels -> internal midpoint category key.
# The norm sheet repeats each midpoint category once per endpoint pathway it
# feeds (e.g. "Global Warming - Human health", "Global Warming - Terrestrial
# ecosystems", ...). The MIDPOINT norm value is the same across those rows
# (kg CO2 eq./person.yr); we collapse to the unique midpoint key.
NORM_MIDPOINT_LABEL_TO_KEY: Dict[str, str] = {
    "Global Warming - Human health":                          "Global warming",
    "Stratospheric ozone depletion - Human health":           "Stratospheric ozone depletion",
    "Ionzing Radiation - Human health":                       "Ionizing radiation",
    "Fine particulate matter formation - Human health":       "Fine particulate matter formation",
    "Photochemical ozone formation - Human health":           "Ozone formation, Human health",
    "Toxicity - Human health (cancer)":                       "Human carcinogenic toxicity",
    "Toxicity - Human health (non-cancer)":                   "Human non-carcinogenic toxicity",
    "Water consumption - human health":                       "Water consumption",
    "Photochemical ozone formation - Terrestrial ecosystems": "Ozone formation, Terrestrial ecosystems",
    "Acidification - Terrestrial ecosystems":                 "Terrestrial acidification",
    "Toxicity - Terrestrial ecosystems":                      "Terrestrial ecotoxicity",
    "Land use - occupation":                                  "Land use",
    "Eutrophication - Freshwater ecosystems":                 "Freshwater eutrophication",
    "Toxicity - Freshwater ecosystems":                       "Freshwater ecotoxicity",
    "Toxicity - Marine ecosystems":                           "Marine ecotoxicity",
    "Eutrophication - marine ecosystems":                     "Marine eutrophication",
    "Mineral resource scarcity":                              "Mineral resource scarcity",
    "Crude oil":                                              "Fossil resource scarcity (crude oil)",
    "Natural gas":                                            "Fossil resource scarcity (natural gas)",
    "Hard coal":                                              "Fossil resource scarcity (hard coal)",
    "Brown coal":                                             "Fossil resource scarcity (brown coal)",
}

# Mapping each endpoint pathway to its area of protection. Used to compute
# per-AoP endpoint normalisation totals from the per-pathway endpoint norms.
ENDPOINT_TO_AOP: Dict[str, str] = {
    # Human Health (DALY)
    "climate_change_human":                  "Human Health",
    "ozone_depletion_human":                 "Human Health",
    "ionizing_radiation_human":              "Human Health",
    "particulate_matter_human":              "Human Health",
    "photochemical_ozone_human":             "Human Health",
    "human_toxicity_cancer":                 "Human Health",
    "human_toxicity_non_cancer":             "Human Health",
    "water_use_human":                       "Human Health",
    # Ecosystems (species.yr)
    "climate_change_ecosystem":              "Ecosystems",
    "photochemical_ozone_ecosystem":         "Ecosystems",
    "terrestrial_acidification_ecosystem":   "Ecosystems",
    "terrestrial_ecotoxicity_ecosystem":     "Ecosystems",
    "water_use_ecosystem_terrestrial":       "Ecosystems",
    "land_use_ecosystem":                    "Ecosystems",
    "climate_change_ecosystem_freshwater":   "Ecosystems",
    "freshwater_eutrophication_ecosystem":   "Ecosystems",
    "freshwater_ecotoxicity_ecosystem":      "Ecosystems",
    "water_use_ecosystem_freshwater":        "Ecosystems",
    "marine_ecotoxicity_ecosystem":          "Ecosystems",
    "marine_eutrophication_ecosystem":       "Ecosystems",
    # Resources (USD2013)
    "mineral_scarcity":                      "Resources",
    "fossil_scarcity_crude_oil":             "Resources",
    "fossil_scarcity_hard_coal":             "Resources",
    "fossil_scarcity_natural_gas":           "Resources",
    "fossil_scarcity_brown_coal":            "Resources",
    "fossil_scarcity_peat":                  "Resources",
}


# ---------------------------------------------------------------------------
# Embedded ISO 3166-1 alpha-3 map for the workbook country strings.
# Covers all 171 entries in the Water consumption sheet plus a few that appear
# only in eutrophication / acidification. Sub-national entries (Canarias,
# Madeira) keep the parent ISO-3 with a region suffix flag.
# ---------------------------------------------------------------------------
COUNTRY_ISO3_MAP: Dict[str, str] = {
    "Afghanistan": "AFG", "Albania": "ALB", "Algeria": "DZA", "Andorra": "AND",
    "Angola": "AGO", "Argentina": "ARG", "Armenia": "ARM", "Australia": "AUS",
    "Austria": "AUT", "Azerbaijan": "AZE", "Bahamas": "BHS", "Bahrain": "BHR",
    "Bangladesh": "BGD", "Barbados": "BRB", "Belarus": "BLR", "Belgium": "BEL",
    "Belize": "BLZ", "Benin": "BEN", "Bhutan": "BTN", "Bolivia": "BOL",
    "Bosnia and Herzegovina": "BIH", "Botswana": "BWA", "Brazil": "BRA",
    "Brunei Darussalam": "BRN", "Bulgaria": "BGR", "Burkina Faso": "BFA",
    "Burundi": "BDI", "Cambodia": "KHM", "Cameroon": "CMR", "Canada": "CAN",
    "Canarias": "ESP-CN", "Cape Verde": "CPV", "Central African Republic": "CAF",
    "Chad": "TCD", "Chile": "CHL", "China": "CHN", "Colombia": "COL",
    "Comoros": "COM", "Congo": "COG", "Congo DRC": "COD", "Costa Rica": "CRI",
    "Côte d'Ivoire": "CIV", "Croatia": "HRV", "Cuba": "CUB", "Cyprus": "CYP",
    "Czech Republic": "CZE", "Denmark": "DNK", "Djibouti": "DJI", "Dominica": "DMA",
    "Dominican Republic": "DOM", "Ecuador": "ECU", "Egypt": "EGY",
    "El Salvador": "SLV", "Equatorial Guinea": "GNQ", "Eritrea": "ERI",
    "Estonia": "EST", "Ethiopia": "ETH", "Fiji": "FJI", "Finland": "FIN",
    "France": "FRA", "French Guiana": "GUF", "Gabon": "GAB", "Gambia": "GMB",
    "Georgia": "GEO", "Germany": "DEU", "Ghana": "GHA", "Greece": "GRC",
    "Greenland": "GRL", "Guatemala": "GTM", "Guinea": "GIN",
    "Guinea-Bissau": "GNB", "Guyana": "GUY", "Haiti": "HTI", "Honduras": "HND",
    "Hong Kong": "HKG", "Hungary": "HUN", "Iceland": "ISL", "India": "IND",
    "Indonesia": "IDN", "Iran": "IRN", "Iraq": "IRQ", "Ireland": "IRL",
    "Israel": "ISR", "Italy": "ITA", "Jamaica": "JAM", "Japan": "JPN",
    "Jordan": "JOR", "Kazakhstan": "KAZ", "Kenya": "KEN", "Kuwait": "KWT",
    "Kyrgyzstan": "KGZ", "Laos": "LAO", "Latvia": "LVA", "Lebanon": "LBN",
    "Lesotho": "LSO", "Liberia": "LBR", "Libya": "LBY", "Liechtenstein": "LIE",
    "Lithuania": "LTU", "Luxembourg": "LUX", "Madagascar": "MDG",
    "Madeira": "PRT-30", "Malawi": "MWI", "Malaysia": "MYS", "Maldives": "MDV",
    "Mali": "MLI", "Malta": "MLT", "Mauritania": "MRT", "Mauritius": "MUS",
    "Mexico": "MEX", "Moldova": "MDA", "Monaco": "MCO", "Mongolia": "MNG",
    "Montenegro": "MNE", "Morocco": "MAR", "Mozambique": "MOZ", "Myanmar": "MMR",
    "Namibia": "NAM", "Nepal": "NPL", "Netherlands": "NLD", "New Zealand": "NZL",
    "Nicaragua": "NIC", "Niger": "NER", "Nigeria": "NGA", "North Korea": "PRK",
    "Norway": "NOR", "Oman": "OMN", "Pakistan": "PAK", "Palestinian Territory": "PSE",
    "Panama": "PAN", "Papua New Guinea": "PNG", "Paraguay": "PRY", "Peru": "PER",
    "Philippines": "PHL", "Poland": "POL", "Portugal": "PRT", "Puerto Rico": "PRI",
    "Qatar": "QAT", "Romania": "ROU", "Russian Federation": "RUS", "Rwanda": "RWA",
    "San Marino": "SMR", "Sao Tome and Principe": "STP", "Saudi Arabia": "SAU",
    "Senegal": "SEN", "Serbia": "SRB", "Sierra Leone": "SLE", "Singapore": "SGP",
    "Slovakia": "SVK", "Slovenia": "SVN", "Somalia": "SOM", "South Africa": "ZAF",
    "South Korea": "KOR", "South Sudan": "SSD", "Spain": "ESP", "Sri Lanka": "LKA",
    "Sudan": "SDN", "Suriname": "SUR", "Swaziland": "SWZ", "Sweden": "SWE",
    "Switzerland": "CHE", "Syria": "SYR", "Taiwan": "TWN", "Tajikistan": "TJK",
    "Tanzania": "TZA", "Thailand": "THA",
    "The Former Yugoslav Republic of Macedonia": "MKD", "Togo": "TGO",
    "Trinidad and Tobago": "TTO", "Tunisia": "TUN", "Turkey": "TUR",
    "Turkmenistan": "TKM", "Uganda": "UGA", "Ukraine": "UKR",
    "United Arab Emirates": "ARE", "United Kingdom": "GBR",
    "United States": "USA", "United States of America": "USA",
    "Uruguay": "URY", "Uzbekistan": "UZB", "Venezuela": "VEN", "Vietnam": "VNM",
    "Yemen": "YEM", "Zambia": "ZMB", "Zimbabwe": "ZWE",
    # Aliases / alternate spellings appearing in other workbook sheets
    "USA": "USA", "UK": "GBR",
    "Antigua and Barbuda": "ATG", "Antigua & Barbuda": "ATG",
    "Bahamas, The": "BHS", "Bahrain": "BHR", "Barbados": "BRB",
    "Bosnia Herzegovina": "BIH", "Brunei": "BRN", "Byelarus": "BLR",
    "Cape Verde": "CPV", "Columbia": "COL",
    "Congo DRC (Zaire)": "COD",
    "Cook Islands": "COK", "Cook island": "COK",
    "Cote d'Ivoire": "CIV", "Cote D'ivoire (Ivory Coast)": "CIV",
    "Ivory Coast": "CIV",
    "East Timor": "TLS", "Timor-Leste": "TLS",
    "Falkland Islands": "FLK", "Falkland Islands (Malvinas)": "FLK",
    "Falkland Islands (Islas Malvinas)": "FLK", "Falkland island": "FLK",
    "Faroe Islands": "FRO", "Faroe Is": "FRO",
    "Federated States of Micronesia": "FSM", "Micronesia": "FSM",
    "France, Metropolitan": "FRA",
    "French Polynesia": "PYF",
    "French Southern & Antarctic Lands": "ATF",
    "French Southern Antarctic Lands": "ATF",
    "Gambia, The": "GMB", "Gaza Strip": "PSE",
    "Gibraltar": "GIB", "Grenada": "GRD", "Guadeloupe": "GLP",
    "Guam": "GUM", "Guernsey": "GGY", "Guinea Bissau": "GNB",
    "Isle of Man": "IMN", "Man, Isle of": "IMN",
    "Jersey": "JEY", "Kiribati": "KIR",
    "Korea, Democratic People's Republic of": "PRK",
    "Korea, Republic of": "KOR",
    "Laos, Peoples Democratic Republic of": "LAO",
    "Liby An Arab Jamahiriya": "LBY",
    "Macau": "MAC", "Macao": "MAC",
    "Macedonia": "MKD", "Macedonia, The Former Republic of Yugoslavia": "MKD",
    "North Macedonia": "MKD",
    "Marshall Islands": "MHL", "Martinique": "MTQ", "Mayotte": "MYT",
    "Moldova, Republic of": "MDA", "Montserrat": "MSR",
    "Myanmar (Burma)": "MMR",
    "Netherlands Antilles": "ANT",
    "New Caledonia": "NCL", "Niue": "NIU", "Norfolk Island": "NFK",
    "Northern Mariana Islands": "MNP", "Northern Marianas": "MNP",
    "Palau": "PLW", "Pitcairn Islands": "PCN", "Pitcairn": "PCN",
    "Reunion": "REU", "Réunion": "REU",
    "Saint Helena": "SHN", "Saint Kitts and Nevis": "KNA",
    "Saint Lucia": "LCA", "Saint Pierre and Miquelon": "SPM",
    "Saint Vincent and The Grenadines": "VCT",
    "Samoa": "WSM", "Western Samoa": "WSM", "American Samoa": "ASM",
    "Seychelles": "SYC", "Solomon Islands": "SLB",
    "Svalbard and Jan Mayen": "SJM",
    "Saint Barthelemy": "BLM", "Saint Martin": "MAF",
    "Sint Maarten": "SXM",
    "Tokelau": "TKL", "Tonga": "TON", "Tuvalu": "TUV",
    "Vanuatu": "VUT", "Vatican City": "VAT", "Holy See": "VAT",
    "Wallis and Futuna": "WLF", "Western Sahara": "ESH",
    "Anguilla": "AIA", "Aruba": "ABW", "Bermuda": "BMU",
    "British Virgin Islands": "VGB", "British Virgin Is": "VGB",
    "Cayman Islands": "CYM", "Christmas Island": "CXR", "Christmas Is": "CXR",
    "Cocos (Keeling) Islands": "CCK",
    "Dominica": "DMA",
    "Heard Island & McDonald Islands": "HMD",
    "British Indian Ocean Territory": "IOT",
    "Antarctica": "ATA", "Bouvet Island": "BVT",
    "Baker Island": "UMI", "Howland Island": "UMI", "Jarvis Island": "UMI",
    "Johnston Atoll": "UMI", "Midway Islands": "UMI",
    "Glorioso Islands": "ATF", "Juan De Nova Island": "ATF",
    # Final batch — remaining workbook variants
    "Jan Mayen": "SJM", "Nauru": "NRU",
    "Netherland": "NLD", "Pacific Islands (Palau)": "PLW",
    "Paracel Islands": "XPI", "Pitcairn island": "PCN",
    "Russia": "RUS", "Sao Tome & Principe": "STP",
    "Serbia & Montenegro": "SRB", "Serbia and Montenegro": "SRB",
    "Solomon Is": "SLB",
    "South Georgia & the South sandwich island": "SGS",
    "South Georgia and the South Sandwich Islands": "SGS",
    "Spratly Islands": "XSP",
    "St Helena": "SHN", "St. Helena": "SHN",
    "St Kitts & Nevis": "KNA", "St. Kitts and Nevis": "KNA",
    "St Lucia": "LCA", "St. Lucia": "LCA",
    "St. Pierre and Miquelon": "SPM",
    "St. Vincent and the Grenadines": "VCT",
    "Svalbard": "SJM", "Syrian Arab Republic": "SYR",
    "Tanzania, United Republic of": "TZA",
    "Trinidad & Tobago": "TTO", "Trinidad Tobago": "TTO",
    "Turks & Caicos Is": "TCA", "Turks and Caicos Islands": "TCA",
    "Virgin Is": "VIR", "Virgin Islands": "VIR",
    "Wake Island": "UMI", "Wallis & Futuna": "WLF",
    "West Bank": "PSE", "Zaire": "COD",
}


def _build_normalized_iso3_map(base_map: Dict[str, str]) -> Dict[str, str]:
    """Lowercase-keyed copy of `base_map` for case-insensitive fallback lookups
    (handles 'Bosnia And Herzegovina' vs 'Bosnia and Herzegovina', etc.)."""
    return {k.lower(): v for k, v in base_map.items()}


_ISO3_MAP_LOWERCASE = _build_normalized_iso3_map(COUNTRY_ISO3_MAP)


# ---------------------------------------------------------------------------
# Helpers (mirrored from build_agribalyse_v32_catalog.py)
# ---------------------------------------------------------------------------

def _normalize_str(value: Any) -> str:
    """Collapse non-breaking spaces, newlines, excess internal whitespace."""
    if value is None:
        return ""
    s = str(value).replace("\xa0", " ").replace("\r", " ").replace("\n", " ")
    return " ".join(s.split())


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:  # NaN
            return None
        return float(value)
    s = str(value).strip()
    if s in ("", "-", "n/a", "N/A", "#N/A", "#DIV/0!", "#VALUE!", "#REF!"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _sha256_of_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _git_rev_short() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL,
        ).decode("ascii").strip()
    except Exception:  # noqa: BLE001
        return "unknown"


# ---------------------------------------------------------------------------
# Extractor 1: midpoint-to-endpoint factors (CF workbook)
# ---------------------------------------------------------------------------

def extract_endpoint_factors_pack(workbook_path: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Parse the 'Midpoint to endpoint factors' sheet.

    Returns:
        pack: {"_schema_version": ..., "perspectives": {"I": {...}, "H": {...}, "E": {...}}}
        meta: {"source_file_sha256": ..., "n_factors_per_perspective": ..., ...}
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if "Midpoint to endpoint factors" not in wb.sheetnames:
        raise ValueError(
            f"'Midpoint to endpoint factors' sheet not found in {workbook_path!r}."
        )
    ws = wb["Midpoint to endpoint factors"]

    factors_by_perspective: Dict[str, Dict[str, float]] = {p: {} for p in PERSPECTIVES}
    unmapped_labels: List[str] = []

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        label = _normalize_str(row[0])
        if not label or label.startswith("Midpoint to endpoint"):
            continue
        # AoP section headers ("Human health", "Terrestrial ecosystems", etc.)
        # have empty value cells; skip cleanly.
        v_i = _coerce_float(row[2]) if len(row) > 2 else None
        v_h = _coerce_float(row[3]) if len(row) > 3 else None
        v_e = _coerce_float(row[4]) if len(row) > 4 else None
        if v_i is None and v_h is None and v_e is None:
            continue
        key = ENDPOINT_LABEL_TO_KEY.get(label)
        if key is None:
            unmapped_labels.append(label)
            continue
        if v_i is not None:
            factors_by_perspective["I"][key] = v_i
        if v_h is not None:
            factors_by_perspective["H"][key] = v_h
        if v_e is not None:
            factors_by_perspective["E"][key] = v_e

    wb.close()

    pack = {
        "_schema_version": SCHEMA_VERSION,
        "_methodology": METHODOLOGY,
        "_methodology_version": METHODOLOGY_VERSION,
        "_source_sheet": "Midpoint to endpoint factors",
        "perspectives": factors_by_perspective,
    }
    meta = {
        "source_file": os.path.basename(workbook_path),
        "source_file_sha256": _sha256_of_file(workbook_path),
        "n_factors_per_perspective": {p: len(d) for p, d in factors_by_perspective.items()},
        "unmapped_labels": unmapped_labels,
    }
    return pack, meta


# ---------------------------------------------------------------------------
# Extractor 2: normalisation scores
# ---------------------------------------------------------------------------

def extract_normalization_pack(workbook_path: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Parse the 'Final normalization scores' sheet.

    The sheet has two side-by-side blocks (Midpoint cols 2-5, Endpoint cols 7-10)
    repeated for each midpoint category per endpoint pathway it feeds.

    Returns:
        pack: {
            "_schema_version": ...,
            "world_population_2010": <int>,
            "midpoint": {"I": {<midpoint_cat>: norm}, "H": {...}, "E": {...}},
            "endpoint_per_pathway": {"I": {<endpoint_key>: norm}, ...},
            "endpoint_per_aop":     {"I": {"Human Health": norm, ...}, ...}
        }
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if "Final normalization scores" not in wb.sheetnames:
        raise ValueError(
            f"'Final normalization scores' sheet not found in {workbook_path!r}."
        )
    ws = wb["Final normalization scores"]

    midpoint_norms: Dict[str, Dict[str, float]] = {p: {} for p in PERSPECTIVES}
    endpoint_pathway_norms: Dict[str, Dict[str, float]] = {p: {} for p in PERSPECTIVES}
    world_population_2010: Optional[int] = None
    unmapped_labels: List[str] = []

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        label = _normalize_str(row[0]) if row[0] is not None else ""

        # Catch the "World population" trailing row
        if label and label.lower().startswith("world population"):
            # Next iteration's row[0] holds the population number (it shows
            # as the cell value below "World population" header per the sheet
            # we sampled). Robust path: try to coerce row[0] to int once.
            continue
        try:
            possible_pop = int(label) if label.isdigit() else None
        except (ValueError, AttributeError):
            possible_pop = None
        if possible_pop and possible_pop > 1_000_000_000:
            world_population_2010 = possible_pop
            continue

        if not label:
            continue
        # Skip AoP section headers (no numeric cells anywhere)
        # Midpoint block: cols 2/3/4 = I/H/E
        m_i = _coerce_float(row[2]) if len(row) > 2 else None
        m_h = _coerce_float(row[3]) if len(row) > 3 else None
        m_e = _coerce_float(row[4]) if len(row) > 4 else None
        # Endpoint block: cols 7/8/9 = I/H/E (col 5 = blank gap, col 6 = unit)
        e_i = _coerce_float(row[7]) if len(row) > 7 else None
        e_h = _coerce_float(row[8]) if len(row) > 8 else None
        e_e = _coerce_float(row[9]) if len(row) > 9 else None
        if all(v is None for v in (m_i, m_h, m_e, e_i, e_h, e_e)):
            continue

        # Midpoint mapping (collapse duplicate rows that repeat the same
        # midpoint norm under different endpoint pathways)
        mid_key = NORM_MIDPOINT_LABEL_TO_KEY.get(label)
        if mid_key:
            if m_i is not None and mid_key not in midpoint_norms["I"]:
                midpoint_norms["I"][mid_key] = m_i
            if m_h is not None and mid_key not in midpoint_norms["H"]:
                midpoint_norms["H"][mid_key] = m_h
            if m_e is not None and mid_key not in midpoint_norms["E"]:
                midpoint_norms["E"][mid_key] = m_e

        # Endpoint pathway mapping
        ep_key = ENDPOINT_LABEL_TO_KEY.get(label)
        if ep_key:
            if e_i is not None:
                endpoint_pathway_norms["I"][ep_key] = e_i
            if e_h is not None:
                endpoint_pathway_norms["H"][ep_key] = e_h
            if e_e is not None:
                endpoint_pathway_norms["E"][ep_key] = e_e

        # Track labels that resolved to NEITHER midpoint nor endpoint
        # (genuine unmapped, not just duplicate-row dedup).
        if mid_key is None and ep_key is None:
            unmapped_labels.append(label)

    wb.close()

    # Compute per-AoP endpoint norms by summing across pathways within each AoP.
    endpoint_per_aop: Dict[str, Dict[str, float]] = {
        p: {"Human Health": 0.0, "Ecosystems": 0.0, "Resources": 0.0} for p in PERSPECTIVES
    }
    for p in PERSPECTIVES:
        for ep_key, val in endpoint_pathway_norms[p].items():
            aop = ENDPOINT_TO_AOP.get(ep_key)
            if aop and val is not None:
                endpoint_per_aop[p][aop] += val

    pack = {
        "_schema_version": SCHEMA_VERSION,
        "_methodology": METHODOLOGY,
        "_methodology_version": METHODOLOGY_VERSION,
        "_source_sheet": "Final normalization scores",
        "world_population_2010": world_population_2010 or 6895889018,
        "midpoint": midpoint_norms,
        "endpoint_per_pathway": endpoint_pathway_norms,
        "endpoint_per_aop": endpoint_per_aop,
    }
    meta = {
        "source_file": os.path.basename(workbook_path),
        "source_file_sha256": _sha256_of_file(workbook_path),
        "n_midpoint_categories": {p: len(d) for p, d in midpoint_norms.items()},
        "n_endpoint_pathways": {p: len(d) for p, d in endpoint_pathway_norms.items()},
        "unmapped_labels": sorted(set(unmapped_labels)),
    }
    return pack, meta


# ---------------------------------------------------------------------------
# Extractor 3: country-specific factors
# ---------------------------------------------------------------------------

def _resolve_iso3(name: str) -> Optional[str]:
    """Look up ISO-3 code from workbook country string. Tries exact match, then
    case-insensitive match; returns None if neither resolves."""
    if not name:
        return None
    stripped = name.strip()
    iso3 = COUNTRY_ISO3_MAP.get(stripped)
    if iso3 is not None:
        return iso3
    return _ISO3_MAP_LOWERCASE.get(stripped.lower())


def _extract_water_consumption_country_block(ws) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    """Water consumption sheet has 3 side-by-side blocks:
      Block A (col 0): Country | E | H | I  (HH endpoint CF, DALY/m3)
      Block B (col 5): Country | I | H | E | all  (terr + aquatic endpoint CF, species.yr/m3)
      Block C (col 11): Country | water requirement ratio (stress index)
    Each block has its OWN country list; we cross-link by ISO-3.
    """
    countries: Dict[str, Dict[str, Any]] = {}
    unmapped: List[str] = []

    for row in ws.iter_rows(values_only=True, min_row=7):
        # Block A: HH endpoint CF
        if row[0]:
            name = _normalize_str(row[0])
            iso3 = _resolve_iso3(name)
            if not iso3:
                unmapped.append(name)
            else:
                e = _coerce_float(row[1]); h = _coerce_float(row[2]); i = _coerce_float(row[3])
                if any(v is not None for v in (e, h, i)):
                    countries.setdefault(iso3, {"_workbook_name": name})
                    countries[iso3]["endpoint_hh"] = {"I": i, "H": h, "E": e}
        # Block B: terrestrial + aquatic endpoint CF
        if len(row) > 5 and row[5]:
            name = _normalize_str(row[5])
            iso3 = _resolve_iso3(name)
            if not iso3:
                unmapped.append(name)
            else:
                i = _coerce_float(row[6]); h = _coerce_float(row[7])
                e = _coerce_float(row[8]); aq = _coerce_float(row[9])
                if any(v is not None for v in (i, h, e, aq)):
                    countries.setdefault(iso3, {"_workbook_name": name})
                    countries[iso3]["endpoint_terrestrial"] = {"I": i, "H": h, "E": e}
                    countries[iso3]["endpoint_aquatic_all_perspectives"] = aq
        # Block C: water stress index
        if len(row) > 11 and row[11]:
            name = _normalize_str(row[11])
            iso3 = _resolve_iso3(name)
            if not iso3:
                unmapped.append(name)
            else:
                stress = _coerce_float(row[12])
                if stress is not None:
                    countries.setdefault(iso3, {"_workbook_name": name})
                    countries[iso3]["water_stress_index"] = stress

    return countries, unmapped


def _extract_freshwater_eutrophication_block(ws) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    """Freshwater eutrophication: single country block.
    Cols (1-indexed): 1=Country, 2=P-to-fresh, 3=PO43--to-fresh, 4=P-to-soil,
    5=PO43--to-soil; cols 8-11 = endpoint species.yr/kg in same order.
    """
    countries: Dict[str, Dict[str, Any]] = {}
    unmapped: List[str] = []
    for row in ws.iter_rows(values_only=True, min_row=7):
        if not row[0]:
            continue
        name = _normalize_str(row[0])
        iso3 = _resolve_iso3(name)
        if not iso3:
            unmapped.append(name)
            continue
        mid = {
            "P_to_freshwater": _coerce_float(row[1]),
            "PO43-_to_freshwater": _coerce_float(row[2]),
            "P_to_soil": _coerce_float(row[3]),
            "PO43-_to_soil": _coerce_float(row[4]),
        }
        endpoint = {
            "P_to_freshwater": _coerce_float(row[7]),
            "PO43-_to_freshwater": _coerce_float(row[8]),
            "P_to_soil": _coerce_float(row[9]),
            "PO43-_to_soil": _coerce_float(row[10]),
        }
        if any(v is not None for v in {**mid, **endpoint}.values()):
            countries[iso3] = {
                "_workbook_name": name,
                "midpoint_cf": {k: v for k, v in mid.items() if v is not None},
                "endpoint_cf_species_yr_per_kg": {k: v for k, v in endpoint.items() if v is not None},
            }
    return countries, unmapped


def _extract_terrestrial_acidification_block(ws) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    """Terrestrial acidification: single country block.
    Cols: 1=Country, 2=NOx-AP, 3=NH3-AP, 4=SO2-AP; cols 6-8 = endpoint species.yr/kg
    """
    countries: Dict[str, Dict[str, Any]] = {}
    unmapped: List[str] = []
    for row in ws.iter_rows(values_only=True, min_row=8):
        if not row[0]:
            continue
        name = _normalize_str(row[0])
        iso3 = _resolve_iso3(name)
        if not iso3:
            unmapped.append(name)
            continue
        midpoint = {
            "NOx": _coerce_float(row[1]),
            "NH3": _coerce_float(row[2]),
            "SO2": _coerce_float(row[3]),
        }
        endpoint = {
            "NOx": _coerce_float(row[5]),
            "NH3": _coerce_float(row[6]),
            "SO2": _coerce_float(row[7]),
        }
        if any(v is not None for v in {**midpoint, **endpoint}.values()):
            countries[iso3] = {
                "_workbook_name": name,
                "midpoint_cf_kgSO2eq_per_kg": {k: v for k, v in midpoint.items() if v is not None},
                "endpoint_cf_species_yr_per_kg": {k: v for k, v in endpoint.items() if v is not None},
            }
    return countries, unmapped


def _extract_region_block(ws, min_row: int) -> Tuple[List[Dict[str, Any]], List[str]]:
    """For PMF and Photochemical ozone formation: rows are SOURCE REGIONS
    (multi-country aggregates like 'Austria, Slovenia, Liechtenstein...').
    We preserve the raw region string + values; downstream code can resolve
    via a country->region lookup. No ISO-3 normalisation attempted here.
    """
    regions: List[Dict[str, Any]] = []
    unmapped: List[str] = []
    for row in ws.iter_rows(values_only=True, min_row=min_row):
        if not row[0]:
            continue
        region_name = _normalize_str(row[0])
        continent = _normalize_str(row[1]) if len(row) > 1 else ""
        values = [_coerce_float(c) for c in row[2:]]
        if any(v is not None for v in values):
            regions.append({
                "source_region": region_name,
                "continent": continent,
                "raw_values": values,
            })
    return regions, unmapped


def extract_country_factors_pack(workbook_path: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Parse all 5 spatially-explicit category sheets from the country workbook."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    categories: Dict[str, Any] = {}
    total_unmapped: List[str] = []

    # 1. Water consumption (clean per-country, 3 blocks)
    if "Water consumption" in wb.sheetnames:
        ws = wb["Water consumption"]
        countries, unmapped = _extract_water_consumption_country_block(ws)
        categories["water_consumption"] = {
            "countries": countries,
            "endpoint_cf_units": {
                "endpoint_hh":           "DALY/m3 consumed",
                "endpoint_terrestrial":  "species.yr/m3 consumed",
                "endpoint_aquatic_all_perspectives": "species.yr/m3 consumed",
                "water_stress_index":    "ratio (0-1)",
            },
            "n_countries": len(countries),
        }
        total_unmapped.extend(unmapped)

    # 2. Freshwater eutrophication (clean per-country)
    if "Freshwater eutrophication" in wb.sheetnames:
        ws = wb["Freshwater eutrophication"]
        countries, unmapped = _extract_freshwater_eutrophication_block(ws)
        categories["freshwater_eutrophication"] = {
            "countries": countries,
            "midpoint_cf_units": "kg P-eq/kg substance",
            "endpoint_cf_units": "species.yr/kg substance",
            "n_countries": len(countries),
        }
        total_unmapped.extend(unmapped)

    # 3. Terrestrial acidification (clean per-country)
    if "Terrestrial acidification" in wb.sheetnames:
        ws = wb["Terrestrial acidification"]
        countries, unmapped = _extract_terrestrial_acidification_block(ws)
        categories["terrestrial_acidification"] = {
            "countries": countries,
            "midpoint_cf_units": "kg SO2-eq/kg substance",
            "endpoint_cf_units": "species.yr/kg substance",
            "n_countries": len(countries),
        }
        total_unmapped.extend(unmapped)

    # 4. Particulate matter formation (REGIONAL aggregates - preserved raw)
    if "Particulate matter formation" in wb.sheetnames:
        ws = wb["Particulate matter formation"]
        regions, unmapped = _extract_region_block(ws, min_row=7)
        categories["particulate_matter_formation"] = {
            "source_regions": regions,
            "note": "Source regions are multi-country aggregates (e.g. 'Austria, Slovenia, Liechtenstein'). Use country->region lookup at runtime.",
            "n_regions": len(regions),
        }
        total_unmapped.extend(unmapped)

    # 5. Photochemical ozone formation (REGIONAL aggregates - preserved raw)
    if "Photochemical ozone formation" in wb.sheetnames:
        ws = wb["Photochemical ozone formation"]
        regions, unmapped = _extract_region_block(ws, min_row=8)
        categories["photochemical_ozone_formation"] = {
            "source_regions": regions,
            "note": "Source regions are multi-country aggregates. Use country->region lookup at runtime.",
            "n_regions": len(regions),
        }
        total_unmapped.extend(unmapped)

    wb.close()

    # Aggregate the union of ISO-3s across the 3 per-country categories so the
    # loader can expose `list_countries()` quickly.
    all_iso3 = set()
    for cat in ("water_consumption", "freshwater_eutrophication", "terrestrial_acidification"):
        if cat in categories:
            all_iso3.update(categories[cat]["countries"].keys())

    pack = {
        "_schema_version": SCHEMA_VERSION,
        "_methodology": METHODOLOGY,
        "_methodology_version": METHODOLOGY_VERSION,
        "_source_workbook": os.path.basename(workbook_path),
        "categories": categories,
        "countries_available_iso3": sorted(all_iso3),
        "_iso3_map_inline": COUNTRY_ISO3_MAP,
    }
    meta = {
        "source_file": os.path.basename(workbook_path),
        "source_file_sha256": _sha256_of_file(workbook_path),
        "n_countries_total": len(all_iso3),
        "unmapped_workbook_names": sorted(set(n for n in total_unmapped if n and n.lower() != "country")),
    }
    return pack, meta


# ---------------------------------------------------------------------------
# Write artefacts
# ---------------------------------------------------------------------------

def _write_pack(pack: Dict[str, Any], out_path: str) -> str:
    """Write the JSON pack in binary mode with deterministic LF line endings
    so the on-disk file matches the SHA-256 computed during write — regardless
    of platform (Windows would otherwise translate LF -> CRLF in text mode)."""
    serialized = json.dumps(pack, ensure_ascii=False, indent=2, sort_keys=False) + "\n"
    blob = serialized.encode("utf-8")
    with open(out_path, "wb") as fh:
        fh.write(blob)
    return hashlib.sha256(blob).hexdigest()


def write_artefacts(
    endpoint_pack: Dict[str, Any],
    endpoint_meta: Dict[str, Any],
    norm_pack: Dict[str, Any],
    norm_meta: Dict[str, Any],
    country_pack: Dict[str, Any],
    country_meta: Dict[str, Any],
    out_dir: str,
) -> Dict[str, str]:
    """Write the 3 packs + combined meta. Returns dict of sha256 sums."""
    os.makedirs(out_dir, exist_ok=True)
    sha_endpoint = _write_pack(endpoint_pack, os.path.join(out_dir, ENDPOINT_PACK_NAME))
    sha_norm = _write_pack(norm_pack, os.path.join(out_dir, NORM_PACK_NAME))
    sha_country = _write_pack(country_pack, os.path.join(out_dir, COUNTRY_PACK_NAME))

    combined_meta = {
        "methodology": METHODOLOGY,
        "methodology_version": METHODOLOGY_VERSION,
        "schema_version": SCHEMA_VERSION,
        "etl_git_rev": _git_rev_short(),
        "extracted_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "packs": {
            "endpoint_factors": {
                "file": ENDPOINT_PACK_NAME,
                "sha256": sha_endpoint,
                **endpoint_meta,
            },
            "normalization": {
                "file": NORM_PACK_NAME,
                "sha256": sha_norm,
                **norm_meta,
            },
            "country_factors": {
                "file": COUNTRY_PACK_NAME,
                "sha256": sha_country,
                **country_meta,
            },
        },
    }
    meta_path = os.path.join(out_dir, META_NAME)
    meta_blob = (json.dumps(combined_meta, ensure_ascii=False, indent=2,
                            sort_keys=True) + "\n").encode("utf-8")
    with open(meta_path, "wb") as fh:
        fh.write(meta_blob)

    # Persist the ISO-3 map alongside the packs so it's reviewable in PRs and
    # the loader can consume it without re-importing the ETL module.
    iso3_path = os.path.join(out_dir, ISO3_MAP_NAME)
    iso3_blob = (json.dumps(
        {"_schema_version": SCHEMA_VERSION, "map": COUNTRY_ISO3_MAP},
        ensure_ascii=False, indent=2, sort_keys=True,
    ) + "\n").encode("utf-8")
    with open(iso3_path, "wb") as fh:
        fh.write(iso3_blob)

    return {
        "endpoint": sha_endpoint,
        "normalization": sha_norm,
        "country": sha_country,
    }


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--data-dir",
        default=DATA_DIR_REL,
        help="Directory holding the 3 workbooks and where outputs are written.",
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Don't write artefacts; print summary only.")
    parser.add_argument("--quiet", action="store_true", help="Suppress info logging.")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    cf_path = os.path.join(args.data_dir, CF_WORKBOOK_NAME)
    norm_path = os.path.join(args.data_dir, NORM_WORKBOOK_NAME)
    country_path = os.path.join(args.data_dir, COUNTRY_WORKBOOK_NAME)

    for p in (cf_path, norm_path, country_path):
        if not os.path.exists(p):
            raise SystemExit(f"Required workbook not found: {p}")

    logger.info("Reading CF workbook: %s", cf_path)
    endpoint_pack, endpoint_meta = extract_endpoint_factors_pack(cf_path)
    logger.info("  -> %d endpoint factors per perspective",
                endpoint_meta["n_factors_per_perspective"]["H"])
    if endpoint_meta["unmapped_labels"]:
        logger.warning("  Unmapped endpoint labels (review): %s",
                       endpoint_meta["unmapped_labels"])

    logger.info("Reading normalisation workbook: %s", norm_path)
    norm_pack, norm_meta = extract_normalization_pack(norm_path)
    logger.info("  -> midpoint norm categories per perspective: %s",
                norm_meta["n_midpoint_categories"])
    logger.info("  -> endpoint norm pathways per perspective: %s",
                norm_meta["n_endpoint_pathways"])
    if norm_meta["unmapped_labels"]:
        logger.info("  (normalisation unmapped labels: %s)",
                    norm_meta["unmapped_labels"])

    logger.info("Reading country-factors workbook: %s", country_path)
    country_pack, country_meta = extract_country_factors_pack(country_path)
    logger.info("  -> %d countries with ISO-3 across per-country categories",
                country_meta["n_countries_total"])
    if country_meta["unmapped_workbook_names"]:
        logger.warning("  Unmapped country names (review): %d names",
                       len(country_meta["unmapped_workbook_names"]))

    if args.dry_run:
        summary = {
            "endpoint_meta": endpoint_meta,
            "norm_meta": norm_meta,
            "country_meta": country_meta,
            "sample_endpoint_H_GW_human": endpoint_pack["perspectives"]["H"].get("climate_change_human"),
            "sample_norm_H_global_warming": norm_pack["midpoint"]["H"].get("Global warming"),
            "sample_water_canada_hh_H": country_pack["categories"].get("water_consumption", {}).get("countries", {}).get("CAN", {}).get("endpoint_hh", {}).get("H"),
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    shas = write_artefacts(
        endpoint_pack, endpoint_meta,
        norm_pack, norm_meta,
        country_pack, country_meta,
        args.data_dir,
    )
    logger.info("Wrote endpoint pack:   %s  (sha256=%s)", ENDPOINT_PACK_NAME, shas["endpoint"][:12])
    logger.info("Wrote normalisation:    %s  (sha256=%s)", NORM_PACK_NAME, shas["normalization"][:12])
    logger.info("Wrote country factors: %s  (sha256=%s)", COUNTRY_PACK_NAME, shas["country"][:12])
    logger.info("Wrote meta:            %s", META_NAME)
    logger.info("Wrote ISO-3 map:       %s", ISO3_MAP_NAME)
    return 0


if __name__ == "__main__":
    sys.exit(main())
