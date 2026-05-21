"""Build the v32 Agribalyse catalog from the published ADEME Tableur workbook.

Idempotent one-shot ETL: input file SHA-256 → deterministic JSON output.
Run from `backend/` after activating the venv:

    python -m environmental_impact_model.etl.build_agribalyse_v32_catalog \\
        --workbook "AGRIBALYSE3.2_Tableur produits alimentaires_PublieAOUT25.xlsx" \\
        --out environmental_impact_model/data/agribalyse_v32_catalog.json

Use `--dry-run` to print row counts + sample rows without writing.

Schema produced (per row, sorted by ciqual_code ascending):
    {
        "ciqual_code":            str (zero-padded if shorter than 5 digits),
        "agb_code":               str,
        "lci_name":               str (English),
        "lci_name_fr":            str (French),
        "agribalyse_group":       str,
        "agribalyse_subgroup":    str,
        "dqr":                    float | None,
        "packaging_approach":     str | None,
        "season_code":            int | None,
        "transport_mode_code":    int | None,
        "delivery":               str | None,
        "preparation":            str | None,
        "recipe2016_midpoints_per_100g": {<ReCiPe key>: float, ...},
        "ef31_indicators_per_100g":      {<EF column header>: float, ...},
        "unit_metadata":          {<EF column header>: <unit string>, ...},
        "warnings":               [<warning tag>, ...],
    }
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import subprocess
import sys
from collections import OrderedDict
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

# Robust import path: support both `python -m environmental_impact_model.etl.build_...`
# (preferred) and direct `python build_agribalyse_v32_catalog.py` invocations.
try:
    from environmental_impact_model.etl.ef_to_recipe_mapping import (
        EF_INCOMPATIBLE_WITH_RECIPE,
        EF_SINGLE_SCORE_COLUMN,
        EF_TO_RECIPE_DIRECT,
        MAPPING_VERSION,
        all_ef_columns,
    )
except ImportError:  # pragma: no cover - fallback for direct invocation
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from ef_to_recipe_mapping import (  # type: ignore[no-redef]
        EF_INCOMPATIBLE_WITH_RECIPE,
        EF_SINGLE_SCORE_COLUMN,
        EF_TO_RECIPE_DIRECT,
        MAPPING_VERSION,
        all_ef_columns,
    )


logger = logging.getLogger(__name__)


# Synthese tab column layout (row 3 contains descriptors + unit strings;
# row 2 contains indicator titles spanning columns 13–32).
SYNTHESE_SHEET_NAME = "Synthese"
HEADER_INDICATOR_ROW = 2   # indicator titles in cols 13-32
HEADER_LABEL_ROW = 3       # descriptor labels in cols 1-12, unit strings in cols 13-32
FIRST_DATA_ROW = 4

# Descriptor columns (1-indexed; openpyxl row tuple is 0-indexed).
DESCRIPTOR_COLUMNS = {
    "agb_code":             1,
    "ciqual_code":          2,
    "agribalyse_group":     3,
    "agribalyse_subgroup":  4,
    "lci_name_fr":          5,
    "lci_name":             6,
    "season_code":          7,
    "transport_mode_code":  8,
    "delivery":             9,
    "packaging_approach":   10,
    "preparation":          11,
    "dqr":                  12,
}

# Indicator columns (13-32). The Synthese title row reports the EF indicator
# names spread across multiple lines — collapse whitespace + nbsp before
# matching against EF_TO_RECIPE_DIRECT / EF_INCOMPATIBLE_WITH_RECIPE.
EXPECTED_INDICATOR_COLUMN_COUNT = 20  # cols 13-32 inclusive

# Per ADEME 2024 errata page (manuscript ref 5; documented in
# code_action_items.md GROUP-D-CODE-1.x-D), these Ciqual codes are flagged
# as withdrawn/suspect and should carry a `warnings` entry so downstream
# consumers can downgrade or refuse them.
ADEME_ERRATA_CIQUAL_CODES = frozenset({
    "26232",  # eggs farm-gate
    "26013",  # Bleu-Blanc-Coeur labelled
    "25998",  # quinoa
    "26037",
    "26034",
    "27029",
    "9901",   # generic placeholder; ADEME confirmed withdrawn
})


def _normalize_str(value: Any) -> str:
    """Collapse non-breaking spaces, embedded newlines / CRs and excess
    internal whitespace; coerce to str. Excel cells frequently carry
    newlines inside wrapped header titles (e.g. 'Code\\nAGB')."""
    if value is None:
        return ""
    s = str(value).replace(" ", " ").replace("\r", " ").replace("\n", " ")
    return " ".join(s.split())


def _normalize_ciqual(value: Any) -> Optional[str]:
    """Coerce a CIQUAL value to a string. Returns None for empty/blank cells."""
    if value is None:
        return None
    if isinstance(value, float) and value != value:  # NaN
        return None
    s = str(value).strip()
    if not s:
        return None
    # Strip a trailing `.0` if openpyxl returned an int as a float.
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _coerce_float(value: Any) -> Optional[float]:
    """Convert an Excel cell value to float. Sentinels (`-`, `n/a`, blank,
    None, error strings) → None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:  # NaN
            return None
        return float(value)
    s = str(value).strip()
    if s in ("", "-", "n/a", "N/A", "#N/A", "#DIV/0!", "#VALUE!", "#REF!"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _coerce_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def _sha256_of_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _git_rev_short() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL,
        ).decode("ascii").strip()
    except Exception:  # noqa: BLE001 - degrade silently
        return "unknown"


def _read_headers(ws) -> Tuple[List[Optional[str]], List[Optional[str]]]:
    """Return (indicator_titles_row2, descriptor_or_unit_labels_row3) as lists
    aligned to 1-indexed columns (index 0 is unused)."""
    indicator_titles: List[Optional[str]] = [None] * (ws.max_column + 1)
    descriptor_labels: List[Optional[str]] = [None] * (ws.max_column + 1)
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), start=1):
        if r_idx == HEADER_INDICATOR_ROW:
            for c_idx, val in enumerate(row, start=1):
                indicator_titles[c_idx] = _normalize_str(val) or None
        elif r_idx == HEADER_LABEL_ROW:
            for c_idx, val in enumerate(row, start=1):
                descriptor_labels[c_idx] = _normalize_str(val) or None
    return indicator_titles, descriptor_labels


def _assert_header_fingerprint(
    indicator_titles: List[Optional[str]],
    descriptor_labels: List[Optional[str]],
) -> None:
    """Abort cleanly if the Synthese header drifts from the locked layout.

    The mapping table assumes specific column positions and indicator labels;
    if ADEME republishes the Tableur with shuffled columns or renamed
    headers, the human must update the mapping. Better to abort than silently
    misalign indicators.
    """
    # Descriptor labels at columns 1-12.
    expected_descriptors = {
        1: "Code AGB",
        2: "Code CIQUAL",
        3: "Groupe d'aliment",
        4: "Sous-groupe d'aliment",
        5: "Nom du Produit en Français",
        6: "LCI Name",
        # Columns 7-12 are long descriptive labels; check just the prefix.
    }
    for col, expected in expected_descriptors.items():
        got = descriptor_labels[col]
        if got is None or not got.startswith(expected):
            raise ValueError(
                f"Synthese header drift: col {col} expected to start with "
                f"{expected!r}, got {got!r}. Update DESCRIPTOR_COLUMNS or the "
                f"workbook layout."
            )
    # Indicator-title row at columns 13-32 must list known EF column names.
    known_ef = all_ef_columns()
    for col in range(13, 33):
        got = indicator_titles[col]
        if got is None:
            raise ValueError(
                f"Synthese header drift: col {col} indicator title missing."
            )
        if got not in known_ef:
            # Fall back to partial-prefix matching (Excel can wrap titles
            # mid-word and `_normalize_str` collapses newlines).
            if not any(got.startswith(k) for k in known_ef):
                raise ValueError(
                    f"Synthese header drift: col {col} indicator {got!r} "
                    f"is not in EF_TO_RECIPE_DIRECT ∪ EF_INCOMPATIBLE_WITH_RECIPE."
                )


def _resolve_indicator_at_col(indicator_titles: List[Optional[str]], col: int) -> Optional[str]:
    """Return the canonical EF column name (one of `all_ef_columns()`) at
    Excel column `col`, or None if it can't be resolved."""
    title = indicator_titles[col]
    if title is None:
        return None
    known = all_ef_columns()
    if title in known:
        return title
    for k in known:
        if title.startswith(k) or k.startswith(title):
            return k
    return None


def extract_catalog(workbook_path: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Read the workbook and produce (sorted_entries, meta).

    Returns:
        entries: list of catalogue rows, sorted ascending by ciqual_code.
        meta:    provenance dict written to the companion `_meta.json`.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if SYNTHESE_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"{SYNTHESE_SHEET_NAME!r} sheet not found in {workbook_path!r}. "
            f"Sheets present: {wb.sheetnames}"
        )
    ws = wb[SYNTHESE_SHEET_NAME]
    indicator_titles, descriptor_labels = _read_headers(ws)
    _assert_header_fingerprint(indicator_titles, descriptor_labels)

    # Build a {col_index: canonical_ef_name} dispatch table once.
    col_to_ef_name: Dict[int, str] = {}
    for col in range(13, 33):
        ef_name = _resolve_indicator_at_col(indicator_titles, col)
        if ef_name is not None:
            col_to_ef_name[col] = ef_name

    # Unit strings live in row 3 cols 13-32 (per `_read_headers` they were
    # captured under descriptor_labels because both share the same row).
    unit_per_col: Dict[int, str] = {
        col: descriptor_labels[col] or "" for col in range(13, 33)
    }

    # Walk data rows (row >= 4). Build raw_entries keyed on Ciqual so we can
    # dedup-keep-last per the plan.
    raw_entries: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()
    duplicate_ciquals_seen: List[str] = []
    rows_with_warnings = 0

    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        if row[DESCRIPTOR_COLUMNS["ciqual_code"] - 1] is None:
            continue
        ciqual = _normalize_ciqual(row[DESCRIPTOR_COLUMNS["ciqual_code"] - 1])
        if ciqual is None:
            continue

        warnings: List[str] = []
        if ciqual in raw_entries:
            duplicate_ciquals_seen.append(ciqual)
            warnings.append("duplicate_ciqual_row_kept_last")
        if ciqual in ADEME_ERRATA_CIQUAL_CODES:
            warnings.append("ademe_errata_flag")

        # Descriptor values.
        agb_code = _normalize_ciqual(row[DESCRIPTOR_COLUMNS["agb_code"] - 1])
        agribalyse_group = _normalize_str(row[DESCRIPTOR_COLUMNS["agribalyse_group"] - 1])
        agribalyse_subgroup = _normalize_str(row[DESCRIPTOR_COLUMNS["agribalyse_subgroup"] - 1])
        lci_name_fr = _normalize_str(row[DESCRIPTOR_COLUMNS["lci_name_fr"] - 1])
        lci_name = _normalize_str(row[DESCRIPTOR_COLUMNS["lci_name"] - 1])
        season_code = _coerce_int(row[DESCRIPTOR_COLUMNS["season_code"] - 1])
        transport_mode_code = _coerce_int(row[DESCRIPTOR_COLUMNS["transport_mode_code"] - 1])
        delivery = _normalize_str(row[DESCRIPTOR_COLUMNS["delivery"] - 1]) or None
        packaging_approach = _normalize_str(row[DESCRIPTOR_COLUMNS["packaging_approach"] - 1]) or None
        preparation = _normalize_str(row[DESCRIPTOR_COLUMNS["preparation"] - 1]) or None
        dqr = _coerce_float(row[DESCRIPTOR_COLUMNS["dqr"] - 1])

        # Indicator values — divide raw per-kg by 10 to convert to per-100g.
        recipe_factors: Dict[str, float] = {}
        ef31_factors: Dict[str, float] = {}
        unit_metadata: Dict[str, str] = {}

        for col, ef_name in col_to_ef_name.items():
            raw = _coerce_float(row[col - 1])
            if raw is None:
                if ef_name == "Changement climatique":
                    warnings.append("missing_climate_change_value")
                continue
            per_100g = raw / 10.0  # ADEME publishes per-kg; pipeline is per-100g.
            ef31_factors[ef_name] = per_100g
            unit_metadata[ef_name] = unit_per_col.get(col, "")
            mapped = EF_TO_RECIPE_DIRECT.get(ef_name)
            if mapped is not None:
                recipe_factors[mapped] = per_100g

        if any(w not in ("duplicate_ciqual_row_kept_last",) for w in warnings):
            rows_with_warnings += 1

        entry: Dict[str, Any] = {
            "ciqual_code": ciqual,
            "agb_code": agb_code,
            "lci_name": lci_name,
            "lci_name_fr": lci_name_fr,
            "agribalyse_group": agribalyse_group,
            "agribalyse_subgroup": agribalyse_subgroup,
            "dqr": dqr,
            "packaging_approach": packaging_approach,
            "season_code": season_code,
            "transport_mode_code": transport_mode_code,
            "delivery": delivery,
            "preparation": preparation,
            "recipe2016_midpoints_per_100g": recipe_factors,
            "ef31_indicators_per_100g": ef31_factors,
            "unit_metadata": unit_metadata,
            "warnings": warnings,
        }
        raw_entries[ciqual] = entry  # dedup-keep-last: later row overwrites earlier.

    entries = sorted(raw_entries.values(), key=lambda e: e["ciqual_code"])

    meta: Dict[str, Any] = {
        "source_file": os.path.basename(workbook_path),
        "source_file_sha256": _sha256_of_file(workbook_path),
        "source_sheet": SYNTHESE_SHEET_NAME,
        "etl_git_rev": _git_rev_short(),
        "mapping_version": MAPPING_VERSION,
        "extracted_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "total_rows": len(entries),
        "rows_with_warnings": rows_with_warnings,
        "duplicate_ciquals_dedup_kept_last": sorted(set(duplicate_ciquals_seen)),
        "ademe_errata_ciqual_codes_flagged": sorted(
            c for c in ADEME_ERRATA_CIQUAL_CODES if c in raw_entries
        ),
    }
    return entries, meta


def write_artefacts(entries: List[Dict[str, Any]], meta: Dict[str, Any], out_path: str) -> str:
    """Write the catalog JSON + sibling meta JSON. Returns the catalog SHA-256."""
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    payload = {
        "_schema_version": "2.0",
        "_provenance_file": os.path.basename(out_path).replace(".json", "_meta.json"),
        "_meta_summary": {
            "total_rows": meta["total_rows"],
            "source_file": meta["source_file"],
            "source_file_sha256": meta["source_file_sha256"],
            "mapping_version": meta["mapping_version"],
        },
        "entries": entries,
    }
    serialized = json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False)
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(serialized)
        fh.write("\n")
    catalog_sha = hashlib.sha256((serialized + "\n").encode("utf-8")).hexdigest()
    meta["catalog_sha256"] = catalog_sha
    meta_path = out_path.replace(".json", "_meta.json")
    with open(meta_path, "w", encoding="utf-8") as fh:
        json.dump(meta, fh, ensure_ascii=False, indent=2, sort_keys=True)
        fh.write("\n")
    return catalog_sha


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, help="Path to the Tableur .xlsx file.")
    parser.add_argument("--out", required=True, help="Path to write the catalog JSON.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Don't write artefacts; print row counts + first 5 entries.")
    parser.add_argument("--quiet", action="store_true", help="Suppress info logging.")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    logger.info("Reading workbook: %s", args.workbook)
    entries, meta = extract_catalog(args.workbook)
    logger.info("Extracted %d rows; %d rows carry warnings; %d duplicate CIQUALs deduped.",
                meta["total_rows"], meta["rows_with_warnings"],
                len(meta["duplicate_ciquals_dedup_kept_last"]))

    if args.dry_run:
        print(json.dumps({
            "meta": meta,
            "sample_first_3": entries[:3],
        }, ensure_ascii=False, indent=2))
        return 0

    sha = write_artefacts(entries, meta, args.out)
    logger.info("Wrote catalog: %s", args.out)
    logger.info("Catalog SHA-256: %s", sha)
    logger.info("Wrote meta:    %s", args.out.replace(".json", "_meta.json"))

    # Best-effort cleanup of the bootstrap embeddings cache so the next
    # matcher load rebuilds embeddings against the new catalog.
    bootstrap_emb = os.path.join(os.path.dirname(args.out), "agribalyse_bootstrap_embeddings.npy")
    if os.path.exists(bootstrap_emb):
        try:
            os.remove(bootstrap_emb)
            logger.info("Removed stale bootstrap embeddings cache: %s", bootstrap_emb)
        except OSError:
            logger.warning("Could not remove stale embeddings cache: %s", bootstrap_emb)

    return 0


if __name__ == "__main__":
    sys.exit(main())
